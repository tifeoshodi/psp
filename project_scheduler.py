"""
Project Scheduler Application
Creates Excel schedules from user input with advanced formatting and calculations.
"""

from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from enum import Enum
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import subprocess
import platform
from dataclasses import dataclass
from enum import Enum
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
import os


class CalendarFormat(Enum):
    FIVE_DAY = "5-day week"
    SIX_DAY = "6-day week"
    SEVEN_DAY = "7-day week"


class ActivitySection(Enum):
    PRE_KICKOFF = "Pre-Kickoff Activities"
    POST_KICKOFF = "Post Kick-off Activities"


@dataclass
class Activity:
    """Represents a single project activity/task"""
    task: str
    action_needed: str
    duration: int
    precursor: str
    sequence: int
    resources: str
    budget: float
    section: ActivitySection

    def __post_init__(self):
        # Ensure duration is positive
        if self.duration < 0:
            self.duration = 0


@dataclass
class Project:
    """Represents the entire project with all activities"""
    title: str
    activities: List[Activity] = field(default_factory=list)
    calendar_format: CalendarFormat = CalendarFormat.FIVE_DAY

    def add_activity(self, activity: Activity):
        """Add an activity to the project"""
        self.activities.append(activity)

    def get_activities_by_section(self, section: ActivitySection) -> List[Activity]:
        """Get all activities in a specific section"""
        return [activity for activity in self.activities if activity.section == section]

    def get_sequences_by_section(self, section: ActivitySection) -> List[int]:
        """Get unique sequences in a section, sorted"""
        activities = self.get_activities_by_section(section)
        sequences = list(set(activity.sequence for activity in activities))
        return sorted(sequences)


class ScheduleCalculator:
    """Handles schedule calculations and calendar adjustments"""

    @staticmethod
    def calculate_schedules(project: Project) -> Dict[int, int]:
        """Calculate schedule values for each sequence"""
        schedules = {}

        # Pre-kickoff activities always have schedule = 0
        pre_kickoff_sequences = project.get_sequences_by_section(ActivitySection.PRE_KICKOFF)
        for seq in pre_kickoff_sequences:
            schedules[seq] = 0

        # Post-kickoff activities have incremental schedules
        post_kickoff_sequences = project.get_sequences_by_section(ActivitySection.POST_KICKOFF)
        post_kickoff_activities = project.get_activities_by_section(ActivitySection.POST_KICKOFF)

        cumulative_schedule = 0
        for seq in post_kickoff_sequences:
            # Find max duration in this sequence
            seq_activities = [a for a in post_kickoff_activities if a.sequence == seq]
            max_duration = max(a.duration for a in seq_activities) if seq_activities else 0

            # Add max duration to cumulative schedule
            cumulative_schedule += max_duration
            schedules[seq] = cumulative_schedule

        return schedules

    @staticmethod
    def apply_calendar_format(schedule_days: int, calendar_format: CalendarFormat) -> int:
        """Apply calendar format adjustments to schedule days"""
        if calendar_format == CalendarFormat.FIVE_DAY:
            # Add 2 days per week for weekends
            weeks = schedule_days // 5
            extra_days = weeks * 2
            return schedule_days + extra_days
        elif calendar_format == CalendarFormat.SIX_DAY:
            # Add 1 day per week for single non-working day
            weeks = schedule_days // 6
            extra_days = weeks * 1
            return schedule_days + extra_days
        else:  # SEVEN_DAY
            return schedule_days

    @staticmethod
    def get_max_duration_activities(project: Project) -> List[Tuple[Activity, int]]:
        """Get activities that have max duration within their sequence group"""
        max_duration_activities = []

        for section in [ActivitySection.PRE_KICKOFF, ActivitySection.POST_KICKOFF]:
            sequences = project.get_sequences_by_section(section)
            section_activities = project.get_activities_by_section(section)

            for seq in sequences:
                seq_activities = [a for a in section_activities if a.sequence == seq]
                if seq_activities:
                    max_duration = max(a.duration for a in seq_activities)
                    for activity in seq_activities:
                        if activity.duration == max_duration:
                            max_duration_activities.append((activity, max_duration))

        return max_duration_activities


class ExcelGenerator:
    """Generates Excel files with proper formatting and calculations"""

    def __init__(self, project: Project):
        self.project = project
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Project Schedule"

        # Styling constants
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True)
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self, output_path: str):
        """Generate the complete Excel file"""
        current_row = 1

        # Add formatted header (project title, logo space, and timestamp)
        current_row = self._add_formatted_header(current_row)
        # No extra row - headers should be consecutive

        # Add column headers
        current_row = self._add_styled_headers(current_row)

        # Add Pre-Kickoff Activities section
        current_row = self._add_section(ActivitySection.PRE_KICKOFF, current_row)
        current_row = self._add_empty_row(current_row)  # Merged empty row

        # Add Post Kick-off Activities section
        current_row = self._add_section(ActivitySection.POST_KICKOFF, current_row)
        
        # Add budget total row
        current_row = self._add_budget_total(current_row)

        # Apply formatting
        self._apply_formatting()

        # Apply sheet protection (allow editing only for Review Comments column)
        self._apply_sheet_protection()

        # Remove gridlines outside content area
        self._remove_external_gridlines()

        # Apply freeze panes for fixed headers
        self._apply_freeze_panes()

        # Generate charts - DISABLED FOR NOW (to be revisited)
        # self._generate_charts(current_row)

        # Save file
        self.workbook.save(output_path)

    def _add_formatted_header(self, start_row: int) -> int:
        """Add formatted header with logo space, project title, and timestamp"""
        import datetime
        import os
        
        # Merge rows 1 & 2 for header
        # Merge cells A-B for logo space and timestamp
        self.worksheet.merge_cells(f"A{start_row}:B{start_row + 1}")
        
        # Add logo and timestamp to A-B cell
        self._add_logo_and_timestamp(start_row)
        
        # Merge cells C-I for project title only (no timestamp)
        self.worksheet.merge_cells(f"C{start_row}:I{start_row + 1}")
        
        # Add project title only
        title_cell = self.worksheet.cell(row=start_row, column=3, value=self.project.title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.fill = PatternFill(start_color="ffc6c6", end_color="ffc6c6", fill_type="solid")
        
        # Set background color for A-B merged cell
        logo_cell = self.worksheet.cell(row=start_row, column=1)
        logo_cell.fill = PatternFill(start_color="ffc6c6", end_color="ffc6c6", fill_type="solid")
        
        # Merge cell J for empty column (no borders)
        self.worksheet.merge_cells(f"J{start_row}:J{start_row + 1}")
        empty_cell = self.worksheet.cell(row=start_row, column=10, value="")
        empty_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        empty_cell.border = None  # Remove border
        
        # Merge cell K for Review Comments space (no borders)
        self.worksheet.merge_cells(f"K{start_row}:K{start_row + 1}")
        review_space_cell = self.worksheet.cell(row=start_row, column=11, value="")
        review_space_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        review_space_cell.border = None  # Remove border
        
        # Set row heights for header rows (halved from doubled)
        self.worksheet.row_dimensions[start_row].height = 35
        self.worksheet.row_dimensions[start_row + 1].height = 35
        
        return start_row + 2  # Skip both merged rows

    def _add_logo_and_timestamp(self, start_row: int):
        """Add logo and timestamp to the merged A-B cell"""
        import os
        import datetime
        
        # Add timestamp text to the merged cell
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        timestamp_text = f"Generated:\n{timestamp}"
        
        timestamp_cell = self.worksheet.cell(row=start_row, column=1, value=timestamp_text)
        timestamp_cell.font = Font(size=9, bold=False)
        timestamp_cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        # Look for logo file in the current directory
        logo_path = "IESL-Logo.png"
        
        if os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image
                
                # Load and resize the image
                img = Image(logo_path)
                
                # Resize image to fit in the top part of the merged cell
                img.width = 50
                img.height = 40
                
                # Position the image in cell A1 (top part of the merged cell)
                img.anchor = f"A{start_row}"
                
                # Add image to worksheet
                self.worksheet.add_image(img)
                
            except Exception as e:
                print(f"Warning: Could not add logo image: {e}")
        else:
            print(f"Warning: Logo file '{logo_path}' not found in current directory")

    def _add_styled_headers(self, start_row: int) -> int:
        """Add styled column headers"""
        headers = [
            "S/No", "Activities/Tasks", "Action Needed", "Duration", "Precursor",
            "Sequence", "Schedule (in days)", "Resources", "Budget (MILLION)", "", "Review Comments"
        ]
        
        # Orange background for columns A-I (Orange, Accent 2, Lighter 80%)
        orange_fill = PatternFill(start_color="FDE4D0", end_color="FDE4D0", fill_type="solid")
        # White background for empty column J
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # Red background for column K (Review Comments)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=start_row, column=col, value=header)
            cell.font = Font(size=11, bold=True, color="000000" if col <= 10 else "FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Apply borders to all columns except empty column J and Review Comments column K
            if col != 10 and col != 11:
                cell.border = self.border
            
            # Apply appropriate background color
            if col <= 9:  # Columns A-I
                cell.fill = orange_fill
            elif col == 10:  # Empty column J
                cell.fill = white_fill
                # No border applied for empty column
            else:  # Column K (Review Comments)
                cell.fill = red_fill
        
        # Set header row height (halved from doubled)
        self.worksheet.row_dimensions[start_row].height = 25
        
        return start_row + 1

    def _add_section(self, section: ActivitySection, start_row: int) -> int:
        """Add a section with its activities"""
        # Determine background color based on section type
        if section == ActivitySection.PRE_KICKOFF:
            section_color = "ffff99"  # Yellow for Pre-kickoff Activities
        else:  # POST_KICKOFF
            section_color = "D2E3A3"  # Olive Green, Accent 3, Lighter 80% for Post kick-off Activities
        
        # Add section header with appropriate background (left-formatted, excluding Review Comments column)
        section_cell = self.worksheet.cell(row=start_row, column=1, value=section.value)
        section_cell.fill = PatternFill(start_color=section_color, end_color=section_color, fill_type="solid")
        section_cell.font = Font(size=12, bold=True)
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        section_cell.border = self.border
        self.worksheet.merge_cells(f"A{start_row}:I{start_row}")  # Merge only to column I, not affecting Review Comments
        
        # Keep column J empty with no border
        j_cell = self.worksheet.cell(row=start_row, column=10, value="")
        j_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        j_cell.border = None
        
        # Keep Review Comments column K separate with no border
        k_cell = self.worksheet.cell(row=start_row, column=11, value="")
        k_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        k_cell.border = None
        
        # Set section header row height (halved from doubled)
        self.worksheet.row_dimensions[start_row].height = 22
        
        current_row = start_row + 1

        # Get activities and schedules for this section
        activities = self.project.get_activities_by_section(section)
        schedules = ScheduleCalculator.calculate_schedules(self.project)
        max_duration_activities = ScheduleCalculator.get_max_duration_activities(self.project)
        max_duration_set = {id(activity) for activity, _ in max_duration_activities}

        # Sort activities by sequence
        activities.sort(key=lambda x: x.sequence)

        # Add activities with S/N numbering
        activity_number = 1
        for activity in activities:
            self._add_activity_row(activity, current_row, schedules, max_duration_set, activity_number)
            current_row += 1
            activity_number += 1

        # Merge cells for same sequences
        self._merge_schedule_cells(section, start_row + 1, current_row - 1)

        return current_row

    def _add_empty_row(self, start_row: int) -> int:
        """Add an empty row with merged columns (excluding Review Comments column)"""
        # Merge columns A to I only (not affecting Review Comments column K)
        self.worksheet.merge_cells(f"A{start_row}:I{start_row}")
        empty_cell = self.worksheet.cell(row=start_row, column=1, value="")
        empty_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Keep column J empty with no border
        j_cell = self.worksheet.cell(row=start_row, column=10, value="")
        j_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        j_cell.border = None
        
        # Keep Review Comments column K separate with no border
        k_cell = self.worksheet.cell(row=start_row, column=11, value="")
        k_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        k_cell.border = None
        
        # Set row height for empty row (back to reasonable size)
        self.worksheet.row_dimensions[start_row].height = 15
        
        return start_row + 1

    def _add_activity_row(self, activity: Activity, row: int, schedules: Dict[int, int], max_duration_set: set, activity_number: int):
        """Add a single activity row with enhanced formatting"""
        # Calculate adjusted schedule - Pre-Kickoff activities always have 0
        if activity.section == ActivitySection.PRE_KICKOFF:
            adjusted_schedule = 0
            schedule_formula = "0"  # Simple value for pre-kickoff
        else:
            raw_schedule = schedules.get(activity.sequence, 0)
            adjusted_schedule = ScheduleCalculator.apply_calendar_format(raw_schedule, self.project.calendar_format)
            
            # Create formula comment for schedule calculation
            if self.project.calendar_format == CalendarFormat.FIVE_DAY:
                schedule_formula = f"={raw_schedule}+({raw_schedule}//5)*2"  # Add weekends
            elif self.project.calendar_format == CalendarFormat.SIX_DAY:
                schedule_formula = f"={raw_schedule}+({raw_schedule}//6)*1"  # Add one day per week
            else:
                schedule_formula = f"={raw_schedule}"  # No adjustment for 7-day

        # Format budget to millions with 2 decimal places
        budget_millions = activity.budget / 1000000 if activity.budget > 0 else 0.0

        values = [
            activity_number,  # S/N column
            activity.task,
            activity.action_needed,
            activity.duration,
            activity.precursor,
            activity.sequence,
            adjusted_schedule,  # We'll add formula separately
            activity.resources,
            budget_millions,
            "",  # Empty column J
            ""   # Empty review comments column K
        ]

        # Determine if this row needs extra height (text-heavy columns) - increase all heights
        text_heavy_columns = [activity.task, activity.action_needed, activity.precursor]
        max_text_length = max(len(str(text)) for text in text_heavy_columns)
        
        # Set row height based on content length (halved from doubled)
        if max_text_length > 50:
            self.worksheet.row_dimensions[row].height = 40
        elif max_text_length > 30:
            self.worksheet.row_dimensions[row].height = 35
        else:
            self.worksheet.row_dimensions[row].height = 30

        # Determine background color for S/No column based on section
        if activity.section == ActivitySection.PRE_KICKOFF:
            sno_bg_color = "ffff99"  # Yellow for Pre-kickoff Activities
        else:  # POST_KICKOFF
            sno_bg_color = "D2E3A3"  # Olive Green, Accent 3, Lighter 80% for Post kick-off Activities

        for col, value in enumerate(values, 1):
            cell = self.worksheet.cell(row=row, column=col, value=value)
            
            # Apply borders consistently (except for empty column J)
            if col != 10:  # All columns except empty column J
                cell.border = self.border
            
            # Apply specific formatting based on column
            if col == 1:  # S/N column with section-specific background color
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color=sno_bg_color, end_color=sno_bg_color, fill_type="solid")
            elif col in [2, 3]:  # Activities/Tasks, Action Needed columns
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.font = Font(size=10)
            elif col in [4, 6, 7]:  # Duration, Sequence, Schedule columns
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)
            elif col in [5, 8]:  # Precursor, Resources columns (center-formatted as requested)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(size=10)
            elif col == 9:  # Budget column  
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.00'  # 2 decimal places
                cell.font = Font(size=10)
            elif col == 10:  # Empty column J
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.border = None  # No border for empty column
            elif col == 11:  # Review Comments column K
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.font = Font(size=10)
                # Unlock this cell for editing
                from openpyxl.styles import Protection
                cell.protection = Protection(locked=False)

            # Highlight max duration text in red
            if col == 4 and id(activity) in max_duration_set:  # Duration column (now col 4)
                cell.font = Font(color="FF0000", bold=True, size=10)  # Red text, bold
        
        # Add formula to schedule cell (column 7) if it's not pre-kickoff
        if activity.section != ActivitySection.PRE_KICKOFF:
            schedule_cell = self.worksheet.cell(row=row, column=7)
            # Add comment with formula explanation
            from openpyxl.comments import Comment
            comment_text = f"Formula: {schedule_formula}\nBase schedule: {schedules.get(activity.sequence, 0)} days\nCalendar format: {self.project.calendar_format.value}"
            comment = Comment(comment_text, "Project Scheduler")
            schedule_cell.comment = comment

    def _merge_schedule_cells(self, section: ActivitySection, start_row: int, end_row: int):
        """Merge cells in Schedule column for same sequence values"""
        activities = self.project.get_activities_by_section(section)
        sequences = self.project.get_sequences_by_section(section)

        for seq in sequences:
            seq_activities = [a for a in activities if a.sequence == seq]
            if len(seq_activities) > 1:
                # Find the rows for this sequence
                seq_rows = []
                current_activities = activities[:]
                current_activities.sort(key=lambda x: x.sequence)

                row_counter = start_row
                for activity in current_activities:
                    if activity.sequence == seq:
                        seq_rows.append(row_counter)
                    row_counter += 1

                if len(seq_rows) > 1:
                    first_row = min(seq_rows)
                    last_row = max(seq_rows)
                    self.worksheet.merge_cells(f"G{first_row}:G{last_row}")  # Schedule is still column G

    def _add_budget_total(self, start_row: int) -> int:
        """Add a total row for budget calculation"""
        current_row = start_row + 1  # Add empty row before total
        
        # Add "Total" label (now in Resources column)
        total_cell = self.worksheet.cell(row=current_row, column=8, value="Total:")
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_cell.border = self.border
        
        # Calculate total budget in millions
        total_budget_millions = sum(activity.budget for activity in self.project.activities) / 1000000
        budget_cell = self.worksheet.cell(row=current_row, column=9, value=total_budget_millions)
        budget_cell.font = Font(bold=True)
        budget_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        budget_cell.alignment = Alignment(horizontal='center', vertical='center')
        budget_cell.number_format = '0.00'  # 2 decimal places
        budget_cell.border = self.border
        
        return current_row + 1

    def _apply_formatting(self):
        """Apply general formatting to the worksheet"""
        # Set specific column widths for better presentation (halved from doubled)
        column_widths = {
            'A': 8,   # S/No - keep thin (halved)
            'B': 30,  # Activities/Tasks - wider for text (halved)
            'C': 30,  # Action Needed - wider for text (halved)
            'D': 12,  # Duration - moderate (halved)
            'E': 25,  # Precursor - wider for text (halved)
            'F': 12,  # Sequence - narrow (halved)
            'G': 18,  # Schedule - moderate (halved)
            'H': 18,  # Resources - moderate (halved)
            'I': 18,  # Budget - moderate (halved)
            'J': 3,   # Empty column - very narrow (halved)
            'K': 25,  # Review Comments - wider for text (halved)
        }
        
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width

        # Apply borders consistently to all content cells
        max_row = self.worksheet.max_row
        max_col = 11  # K column (Review Comments)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                # Skip borders for empty column J only
                if col == 10:  # Empty column J - never has borders
                    continue
                # Apply borders to all other cells (including Review Comments column)
                if cell.value is not None or (row >= 1 and col <= max_col and col != 10):
                    if not hasattr(cell, 'border') or cell.border is None:  # Only apply if not already set
                        cell.border = self.border

    def _apply_sheet_protection(self):
        """Apply sheet protection, leaving only Review Comments column unlocked"""
        from openpyxl.styles import Protection
        
        # Lock all cells by default
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.protection = Protection(locked=True)
        
        # Unlock Review Comments column (column K, which is column 11)
        for row in self.worksheet.iter_rows(min_col=11, max_col=11):
            for cell in row:
                # Only unlock data rows (skip headers and section headers)
                if cell.row > 3:  # Assuming headers start around row 3
                    cell.protection = Protection(locked=False)
        
        # Protect the worksheet with a password (optional)
        # You can change or remove the password as needed
        self.worksheet.protection.sheet = True
        self.worksheet.protection.enable()

    def _remove_external_gridlines(self):
        """Remove gridlines outside the content area to make it plain white"""
        # Hide gridlines for the entire worksheet
        self.worksheet.sheet_view.showGridLines = False

    def _apply_freeze_panes(self):
        """Apply freeze panes to keep headers fixed while scrolling"""
        # Freeze the first 3 rows (header rows 1-2 + column headers row 3)
        # This will keep them visible when scrolling
        self.worksheet.freeze_panes = 'A4'  # Freeze everything above row 4

    # CHART GENERATION METHODS - DISABLED FOR NOW (to be revisited)
    
    # def _generate_charts(self, start_row: int):
    #     """Generate Gantt chart and S-curve on the same sheet"""
    #     try:
    #         # Calculate chart data (no date dependency needed)
    #         chart_data = self._prepare_chart_data()
    #         
    #         # Create new worksheet for chart data tables
    #         self.chart_worksheet = self.workbook.create_sheet("Chart Data")
    #         
    #         # Add title to chart data worksheet
    #         title_cell = self.chart_worksheet.cell(row=1, column=1, value="Project Chart Data Tables")
    #         title_cell.font = Font(size=16, bold=True)
    #         self.chart_worksheet.merge_cells("A1:E1")
    #         
    #         # Set column widths for chart data worksheet
    #         chart_column_widths = {
    #             'A': 40,  # Activity names
    #             'B': 15,  # Start Time
    #             'C': 12,  # Duration
    #             'D': 25,  # Section
    #             'E': 15,  # Critical Path
    #         }
    #         
    #         for col_letter, width in chart_column_widths.items():
    #             self.chart_worksheet.column_dimensions[col_letter].width = width
    #         
    #         # Position charts below the main schedule table (no data tables on main sheet)
    #         chart_start_row = start_row + 2
    #         
    #         # Generate Gantt chart
    #         gantt_end_row = self._create_gantt_chart(chart_data, chart_start_row)
    #         
    #         # Generate S-curve below Gantt chart
    #         s_curve_start_row = gantt_end_row + 2
    #         self._create_s_curve(chart_data, s_curve_start_row)
    #         
    #     except Exception as e:
    #         print(f"Warning: Could not generate charts: {e}")

    # def _prepare_chart_data(self):
    #     """Prepare data needed for both Gantt chart and S-curve"""
    #     # Calculate schedules for all activities
    #     schedules = ScheduleCalculator.calculate_schedules(self.project)
    #     
    #     # Get all activities sorted by section and sequence
    #     all_activities = []
    #     for section in [ActivitySection.PRE_KICKOFF, ActivitySection.POST_KICKOFF]:
    #         activities = self.project.get_activities_by_section(section)
    #         activities.sort(key=lambda x: x.sequence)
    #         all_activities.extend(activities)
    #     
    #     # Calculate critical path using existing max duration logic
    #     max_duration_activities = ScheduleCalculator.get_max_duration_activities(self.project)
    #     critical_path_activities = {id(activity) for activity, _ in max_duration_activities}
    #     
    #     chart_data = {
    #         'activities': [],
    #         'progress_timeline': []
    #     }
    #     
    #     # Process each activity for Gantt chart
    #     for activity in all_activities:
    #         # Calculate schedule position (start time in project timeline)
    #         if activity.section == ActivitySection.PRE_KICKOFF:
    #             # Pre-kickoff activities have negative start times
    #             start_time = -schedules.get(activity.sequence, 0)
    #             end_time = start_time + activity.duration
    #         else:
    #             # Post-kickoff activities start at their scheduled time
    #             start_time = schedules.get(activity.sequence, 0)
    #             end_time = start_time + activity.duration
    #         
    #         # Check if activity is on critical path
    #         is_critical = id(activity) in critical_path_activities
    #         
    #         # Add to activities data
    #         chart_data['activities'].append({
    #             'name': activity.task,
    #             'start_time': start_time,
    #             'end_time': end_time,
    #             'duration': activity.duration,
    #             'budget': activity.budget,
    #             'section': activity.section.value,
    #             'sequence': activity.sequence,
    #             'is_critical': is_critical
    #         })
    #     
    #     # Create ideal progress timeline for S-curve
    #     # Sort activities by end time to create cumulative progress
    #     sorted_activities = sorted(chart_data['activities'], key=lambda x: x['end_time'])
    #     
    #     total_activities = len(sorted_activities)
    #     total_duration = sum(act['duration'] for act in sorted_activities)
    #     
    #     cumulative_progress = 0
    #     cumulative_duration = 0
    #     
    #     for i, activity in enumerate(sorted_activities):
    #         cumulative_progress = ((i + 1) / total_activities) * 100  # Activity completion percentage
    #         cumulative_duration += activity['duration']
    #         duration_progress = (cumulative_duration / total_duration) * 100 if total_duration > 0 else 0
    #         
    #         chart_data['progress_timeline'].append({
    #             'time': activity['end_time'],
    #             'activity_progress': cumulative_progress,
    #             'duration_progress': duration_progress,
    #             'cumulative_duration': cumulative_duration
    #         })
    #     
    #     return chart_data



    # def _create_gantt_chart(self, chart_data, start_row: int) -> int:
    #     """Create a Gantt chart showing all activities with critical path highlighted"""
    #     try:
    #         from openpyxl.chart import BarChart, Reference, Series
    #         
    #         # Add title on main sheet
    #         title_cell = self.worksheet.cell(row=start_row, column=1, value="Project Gantt Chart - All Activities")
    #         title_cell.font = Font(size=14, bold=True)
    #         
    #         # Create data table on chart worksheet
    #         self._create_gantt_data_table(chart_data)
    #         
    #         # Create chart on main sheet but reference data from chart worksheet
    #         chart = BarChart()
    #         chart.type = "bar"
    #         chart.style = 12
    #         chart.title = "Project Timeline - Critical Path (Red) vs Normal Activities (Grey)"
    #         chart.y_axis.title = "Activities"
    #         chart.x_axis.title = "Duration (Days)"
    #         
    #         # Reference data from chart worksheet (starting from row 5, after headers)
    #         duration_data = Reference(self.chart_worksheet, min_col=3, min_row=5,
    #                                 max_col=3, max_row=4 + len(chart_data['activities']))
    #         cats = Reference(self.chart_worksheet, min_col=1, min_row=5,
    #                        max_col=1, max_row=4 + len(chart_data['activities']))
    #         
    #         # Add duration data as bars
    #         chart.add_data(duration_data, titles_from_data=False)
    #         chart.set_categories(cats)
    #         
    #         # Style the chart with conditional colors
    #         if chart.series:
    #             s1 = chart.series[0]
    #             # Set default grey color for all bars
    #             s1.graphicalProperties.solidFill = "808080"  # Grey for normal activities
    #             
    #             # Note: Individual bar coloring for critical path (red) would require 
    #             # more complex chart manipulation or separate series
    #         
    #         # Position the chart
    #         chart.anchor = f"A{start_row + 2}"
    #         chart.width = 25
    #         chart.height = max(15, len(chart_data['activities']) * 0.4)  # Scale height with number of activities
    #         
    #         self.worksheet.add_chart(chart)
    #         
    #         return start_row + int(chart.height / 15) + 3  # Estimate end row based on chart height
    #         
    #     except Exception as e:
    #         print(f"Warning: Could not create Gantt chart: {e}")
    #         return start_row + 10

    # def _create_gantt_data_table(self, chart_data):
    #     """Create Gantt chart data table on the chart worksheet"""
    #     # Start after title
    #     gantt_title_row = 3
    #     gantt_header_row = 4
    #     
    #     # Add Gantt section title
    #     gantt_title_cell = self.chart_worksheet.cell(row=gantt_title_row, column=1, value="Gantt Chart Data")
    #     gantt_title_cell.font = Font(size=12, bold=True)
    #     
    #     headers = ["Activity", "Start Time", "Duration", "Section", "Critical Path"]
    #     
    #     # Add headers
    #     for col, header in enumerate(headers, 1):
    #         cell = self.chart_worksheet.cell(row=gantt_header_row, column=col, value=header)
    #         cell.font = Font(bold=True)
    #         cell.border = self.border
    #         if col == 5:  # Critical Path column
    #             cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    #     
    #     # Add all activity data
    #     for row, activity in enumerate(chart_data['activities'], gantt_header_row + 1):
    #         # Activity name
    #         name_cell = self.chart_worksheet.cell(row=row, column=1, value=activity['name'])
    #         name_cell.border = self.border
    #         
    #         # Start time
    #         start_cell = self.chart_worksheet.cell(row=row, column=2, value=activity['start_time'])
    #         start_cell.border = self.border
    #         
    #         # Duration
    #         duration_cell = self.chart_worksheet.cell(row=row, column=3, value=activity['duration'])
    #         duration_cell.border = self.border
    #         
    #         # Section
    #         section_cell = self.chart_worksheet.cell(row=row, column=4, value=activity['section'])
    #         section_cell.border = self.border
    #         
    #         # Critical path indicator
    #         critical_cell = self.chart_worksheet.cell(row=row, column=5, value="YES" if activity['is_critical'] else "NO")
    #         critical_cell.border = self.border
    #         if activity['is_critical']:
    #             critical_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    #             critical_cell.font = Font(color="FFFFFF", bold=True)
    #             # Also highlight the activity name in red
    #             name_cell.font = Font(color="FF0000", bold=True)

    # def _create_s_curve(self, chart_data, start_row: int) -> int:
    #     """Create an S-curve showing ideal project progress over time"""
    #     try:
    #         from openpyxl.chart import LineChart, Reference
    #         
    #         # Add title on main sheet
    #         title_cell = self.worksheet.cell(row=start_row, column=1, value="Project S-Curve - Ideal Progress Over Time")
    #         title_cell.font = Font(size=14, bold=True)
    #         
    #         # Create data table on chart worksheet
    #         self._create_s_curve_data_table(chart_data)
    #         
    #         # Create chart on main sheet but reference data from chart worksheet
    #         chart = LineChart()
    #         chart.title = "Project S-Curve - Ideal Progress Timeline"
    #         chart.style = 13
    #         chart.y_axis.title = "Progress Percentage (%)"
    #         chart.x_axis.title = "Project Timeline (Days)"
    #         
    #         # Find the start row for S-curve data on chart worksheet (after Gantt data)
    #         s_curve_start_row = 4 + len(chart_data['activities']) + 3  # Gap after Gantt data
    #         
    #         # Reference data from chart worksheet
    #         activity_data = Reference(self.chart_worksheet, min_col=2, min_row=s_curve_start_row,
    #                                 max_col=2, max_row=s_curve_start_row + len(chart_data['progress_timeline']))
    #         duration_data = Reference(self.chart_worksheet, min_col=3, min_row=s_curve_start_row,
    #                                 max_col=3, max_row=s_curve_start_row + len(chart_data['progress_timeline']))
    #         timeline = Reference(self.chart_worksheet, min_col=1, min_row=s_curve_start_row + 1,
    #                            max_col=1, max_row=s_curve_start_row + len(chart_data['progress_timeline']))
    #         
    #         # Add both progress curves
    #         chart.add_data(activity_data, titles_from_data=True)
    #         chart.add_data(duration_data, titles_from_data=True)
    #         chart.set_categories(timeline)
    #         
    #         # Style the lines
    #         if len(chart.series) >= 1:
    #             s1 = chart.series[0]  # Activity completion progress
    #             s1.marker.symbol = "circle"
    #             s1.marker.size = 4
    #             s1.smooth = True
    #             s1.graphicalProperties.line.solidFill = "0066CC"  # Blue
    #         
    #         if len(chart.series) >= 2:
    #             s2 = chart.series[1]  # Duration progress
    #             s2.marker.symbol = "triangle"
    #             s2.marker.size = 4
    #             s2.smooth = True
    #             s2.graphicalProperties.line.solidFill = "FF6600"  # Orange
    #         
    #         # Position the chart
    #         chart.anchor = f"A{start_row + 2}"
    #         chart.width = 25
    #         chart.height = 12
    #         
    #         self.worksheet.add_chart(chart)
    #         
    #         return start_row + 15
    #         
    #     except Exception as e:
    #         print(f"Warning: Could not create S-curve: {e}")
    #         return start_row + 10

    # def _create_s_curve_data_table(self, chart_data):
    #     """Create S-curve data table on the chart worksheet"""
    #     # Start after Gantt data
    #     s_curve_title_row = 4 + len(chart_data['activities']) + 2
    #     start_row = 4 + len(chart_data['activities']) + 3
    #     
    #     # Add S-curve section title
    #     s_curve_title_cell = self.chart_worksheet.cell(row=s_curve_title_row, column=1, value="S-Curve Data")
    #     s_curve_title_cell.font = Font(size=12, bold=True)
    #     
    #     headers = ["Timeline (Days)", "Activity Progress (%)", "Duration Progress (%)", "Cumulative Duration"]
    #     
    #     # Add headers
    #     for col, header in enumerate(headers, 1):
    #         cell = self.chart_worksheet.cell(row=start_row, column=col, value=header)
    #         cell.font = Font(bold=True)
    #         cell.border = self.border
    #         if col == 2:  # Activity Progress column
    #             cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    #         elif col == 3:  # Duration Progress column
    #             cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    #     
    #     # Add progress timeline data
    #     for row, item in enumerate(chart_data['progress_timeline'], start_row + 1):
    #         # Timeline (in days)
    #         time_cell = self.chart_worksheet.cell(row=row, column=1, value=item['time'])
    #         time_cell.border = self.border
    #         
    #         # Activity progress percentage
    #         activity_cell = self.chart_worksheet.cell(row=row, column=2, value=round(item['activity_progress'], 1))
    #         activity_cell.border = self.border
    #         activity_cell.number_format = '0.0'
    #         
    #         # Duration progress percentage
    #         duration_cell = self.chart_worksheet.cell(row=row, column=3, value=round(item['duration_progress'], 1))
    #         duration_cell.border = self.border
    #         duration_cell.number_format = '0.0'
    #         
    #         # Cumulative duration
    #         cumulative_cell = self.chart_worksheet.cell(row=row, column=4, value=item['cumulative_duration'])
         #         cumulative_cell.border = self.border

    # END OF CHART GENERATION METHODS - ALL DISABLED


class ExcelLoader:
    """Loads existing Excel files generated by the project scheduler"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def load_project(self) -> Optional[Project]:
        """Load project data from Excel file"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            # Extract project data
            project_title = self._extract_project_title()
            calendar_format = self._extract_calendar_format()
            activities = self._extract_activities()
            
            if not project_title:
                raise ValueError("Could not find project title in the Excel file")
            
            return Project(
                title=project_title,
                calendar_format=calendar_format,
                activities=activities
            )
            
        except Exception as e:
            raise Exception(f"Error loading Excel file: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _extract_project_title(self) -> Optional[str]:
        """Extract project title from the first merged cell"""
        for row in range(1, 10):  # Check first 10 rows
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str) and "Project Title:" in cell.value:
                return cell.value.replace("Project Title:", "").strip()
        return None
    
    def _extract_calendar_format(self) -> CalendarFormat:
        """Extract calendar format from schedule calculations or default to 6-day"""
        # For now, default to 6-day format
        # Could be enhanced to detect from formulas in comments
        return CalendarFormat.SIX_DAY
    
    def _extract_activities(self) -> List[Activity]:
        """Extract activities from the Excel sheet"""
        activities = []
        current_section = None
        header_row = None
        
        # Find the header row
        for row in range(1, 20):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value == "S/N":
                header_row = row
                break
        
        if not header_row:
            raise ValueError("Could not find header row with S/N column")
        
        # Process rows after header
        for row in range(header_row + 1, self.worksheet.max_row + 1):
            row_data = self._get_row_data(row)
            
            if not any(row_data):  # Skip empty rows
                continue
            
            # Check if this is a section header
            if self._is_section_row(row):
                section_name = self.worksheet.cell(row=row, column=1).value
                if "Pre-Kickoff" in section_name or "Pre Kickoff" in section_name:
                    current_section = ActivitySection.PRE_KICKOFF
                elif "Post Kick-off" in section_name or "Post Kickoff" in section_name:
                    current_section = ActivitySection.POST_KICKOFF
                continue
            
            # Check if this is the total row
            if self._is_total_row(row):
                break
            
            # Extract activity data
            try:
                s_n = row_data[0] if row_data[0] else ""
                task = row_data[1] if row_data[1] else ""
                action_needed = row_data[2] if row_data[2] else ""
                duration = int(row_data[3]) if row_data[3] and str(row_data[3]).replace('.', '').isdigit() else 0
                precursor = row_data[4] if row_data[4] else ""
                sequence = int(row_data[5]) if row_data[5] and str(row_data[5]).replace('.', '').isdigit() else 0
                resources = row_data[7] if row_data[7] else ""
                
                # Convert budget from millions back to base currency
                budget_millions = float(row_data[8]) if row_data[8] and str(row_data[8]).replace('.', '').replace('-', '').isdigit() else 0.0
                budget = budget_millions * 1000000
                
                if task and current_section:  # Only add if we have a task and section
                    activity = Activity(
                        task=task,
                        action_needed=action_needed,
                        duration=duration,
                        precursor=precursor,
                        sequence=sequence,
                        resources=resources,
                        budget=budget,
                        section=current_section
                    )
                    activities.append(activity)
                    
            except (ValueError, TypeError) as e:
                # Skip rows with invalid data
                continue
        
        return activities
    
    def _get_row_data(self, row: int) -> List:
        """Get all cell values from a row"""
        return [self.worksheet.cell(row=row, column=col).value for col in range(1, 10)]
    
    def _is_section_row(self, row: int) -> bool:
        """Check if row is a section header"""
        cell = self.worksheet.cell(row=row, column=1)
        if not cell.value:
            return False
        
        value = str(cell.value).lower()
        return ("kickoff" in value or "kick-off" in value) and self._is_merged_across_columns(row, 1)
    
    def _is_total_row(self, row: int) -> bool:
        """Check if row is the total budget row"""
        # Check if "Total:" appears in Resources column (column 8)
        total_cell = self.worksheet.cell(row=row, column=8)
        return total_cell.value and str(total_cell.value).strip().lower() == "total:"
    
    def _is_merged_across_columns(self, row: int, col: int) -> bool:
        """Check if cell is merged across multiple columns"""
        cell = self.worksheet.cell(row=row, column=col)
        for merged_range in self.worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False


class ProjectSchedulerGUI:
    """Main GUI application for project scheduling"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Project Scheduler")
        self.root.geometry("1200x800")  # Increased size to accommodate all elements
        self.root.minsize(1000, 700)  # Set minimum size
        
        self.project = None
        self.activities_data = []

        self.setup_ui()

    def setup_ui(self):
        """Set up the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)  # Make activities section expandable

        # Project title section
        self.setup_project_section(main_frame)

        # Calendar format section
        self.setup_calendar_section(main_frame)

        # Activities section
        self.setup_activities_section(main_frame)

        # Buttons section
        self.setup_buttons_section(main_frame)

    def setup_project_section(self, parent):
        """Set up project title input section"""
        # Project Title
        ttk.Label(parent, text="Project Title:", font=('Arial', 12, 'bold')).grid(
            row=0, column=0, sticky=tk.W, pady=(0, 10))

        self.project_title_var = tk.StringVar()
        project_entry = ttk.Entry(parent, textvariable=self.project_title_var, width=50, font=('Arial', 11))
        project_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 10))

    def setup_calendar_section(self, parent):
        """Set up calendar format selection"""
        ttk.Label(parent, text="Calendar Format:", font=('Arial', 12, 'bold')).grid(
            row=1, column=0, sticky=tk.W, pady=(0, 10))

        self.calendar_format_var = tk.StringVar(value=CalendarFormat.FIVE_DAY.value)
        calendar_frame = ttk.Frame(parent)
        calendar_frame.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(0, 10))

        for i, fmt in enumerate(CalendarFormat):
            ttk.Radiobutton(calendar_frame, text=fmt.value, variable=self.calendar_format_var,
                          value=fmt.value).grid(row=0, column=i, padx=(0, 20), sticky=tk.W)

    def setup_activities_section(self, parent):
        """Set up activities input section"""
        # Activities section label
        ttk.Label(parent, text="Activities:", font=('Arial', 12, 'bold')).grid(
            row=2, column=0, columnspan=2, sticky=tk.W, pady=(20, 10))

        # Activities frame
        activities_frame = ttk.Frame(parent)
        activities_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 20))
        activities_frame.columnconfigure(0, weight=1)
        activities_frame.rowconfigure(1, weight=1)

        # Activities input form
        self.setup_activity_form(activities_frame)

        # Activities list
        self.setup_activities_list(activities_frame)

    def setup_activity_form(self, parent):
        """Set up the form for adding new activities"""
        form_frame = ttk.LabelFrame(parent, text="Add New Activity", padding="10")
        form_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        form_frame.columnconfigure(1, weight=1)
        form_frame.columnconfigure(3, weight=1)

        # Form fields
        fields = [
            ("Task:", "task_var"),
            ("Action Needed:", "action_var"),
            ("Duration (days):", "duration_var"),
            ("Precursor:", "precursor_var"),
            ("Sequence:", "sequence_var"),
            ("Resources:", "resources_var"),
            ("Budget:", "budget_var")
        ]

        self.form_vars = {}
        for i, (label, var_name) in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2

            ttk.Label(form_frame, text=label).grid(row=row, column=col, sticky=tk.W, padx=(0, 5), pady=2)

            if var_name in ['duration_var', 'sequence_var', 'budget_var']:
                self.form_vars[var_name] = tk.StringVar()
                entry = ttk.Entry(form_frame, textvariable=self.form_vars[var_name], width=20)
            else:
                self.form_vars[var_name] = tk.StringVar()
                entry = ttk.Entry(form_frame, textvariable=self.form_vars[var_name], width=30)

            entry.grid(row=row, column=col+1, sticky=(tk.W, tk.E), padx=(0, 20), pady=2)

        # Section selection
        ttk.Label(form_frame, text="Section:").grid(row=4, column=0, sticky=tk.W, padx=(0, 5), pady=2)
        self.section_var = tk.StringVar(value=ActivitySection.PRE_KICKOFF.value)
        section_combo = ttk.Combobox(form_frame, textvariable=self.section_var,
                                   values=[s.value for s in ActivitySection], state="readonly", width=25)
        section_combo.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(0, 20), pady=2)

        # Add button
        ttk.Button(form_frame, text="Add Activity", command=self.add_activity).grid(
            row=4, column=3, padx=(20, 0), pady=10)

    def setup_activities_list(self, parent):
        """Set up the list view for activities"""
        self.list_frame = ttk.LabelFrame(parent, text="Activities List (0 activities)", padding="10")
        self.list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.list_frame.columnconfigure(0, weight=1)
        self.list_frame.rowconfigure(0, weight=1)

        # Treeview for activities (updated to show budget in millions)
        columns = ("Section", "Task", "Action", "Duration", "Precursor", "Sequence", "Resources", "Budget (M)")
        self.activities_tree = ttk.Treeview(self.list_frame, columns=columns, show="headings", height=15)

        # Configure columns
        for col in columns:
            self.activities_tree.heading(col, text=col)
            self.activities_tree.column(col, width=100, minwidth=80)

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(self.list_frame, orient=tk.VERTICAL, command=self.activities_tree.yview)
        h_scrollbar = ttk.Scrollbar(self.list_frame, orient=tk.HORIZONTAL, command=self.activities_tree.xview)
        self.activities_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Grid layout
        self.activities_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))

        # Delete and Edit buttons and right-click context menu
        delete_frame = ttk.Frame(self.list_frame)
        delete_frame.grid(row=2, column=0, columnspan=2, pady=(10, 0), sticky=(tk.W, tk.E))
        
        ttk.Button(delete_frame, text="Delete Selected", command=self.delete_activity_with_confirmation).pack(
            side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(delete_frame, text="Edit Selected", command=self.edit_activity_with_dialog).pack(
            side=tk.LEFT, padx=(0, 10))
        
        # Setup right-click context menu
        self.setup_context_menu()

    def setup_buttons_section(self, main_frame):
        """Setup the buttons section"""
        buttons_frame = ttk.LabelFrame(main_frame, text="Actions", padding="10")
        buttons_frame.grid(row=4, column=0, columnspan=2, pady=(20, 10), sticky=(tk.W, tk.E))
        
        # Create a centered button layout
        button_container = ttk.Frame(buttons_frame)
        button_container.pack(expand=True)
        
        # Load Excel button
        self.load_button = ttk.Button(
            button_container, 
            text="Load Excel file", 
            command=self.load_excel,
            width=15
        )
        self.load_button.pack(side=tk.LEFT, padx=5)
        
        # Add Activity button
        self.add_button = ttk.Button(
            button_container, 
            text="Add Activity", 
            command=self.add_activity,
            width=15
        )
        self.add_button.pack(side=tk.LEFT, padx=5)
        
        # Preview button
        self.preview_button = ttk.Button(
            button_container, 
            text="Preview", 
            command=self.preview_schedule,
            width=15
        )
        self.preview_button.pack(side=tk.LEFT, padx=5)
        
        # Generate Excel button
        self.generate_button = ttk.Button(
            button_container, 
            text="Generate Excel", 
            command=self.generate_excel,
            width=15
        )
        self.generate_button.pack(side=tk.LEFT, padx=5)
        
        # Clear All button
        self.clear_button = ttk.Button(
            button_container, 
            text="Clear All", 
            command=self.clear_all_with_confirmation,
            width=15
        )
        self.clear_button.pack(side=tk.LEFT, padx=5)

    def add_activity(self):
        """Add a new activity to the list"""
        try:
            # Validate inputs
            task = self.form_vars['task_var'].get().strip()
            action = self.form_vars['action_var'].get().strip()
            duration_str = self.form_vars['duration_var'].get().strip()
            precursor = self.form_vars['precursor_var'].get().strip()
            sequence_str = self.form_vars['sequence_var'].get().strip()
            resources = self.form_vars['resources_var'].get().strip()
            budget_str = self.form_vars['budget_var'].get().strip()
            section_str = self.section_var.get()

            if not all([task, action, duration_str, sequence_str]):
                messagebox.showerror("Error", "Please fill in Task, Action Needed, Duration, and Sequence fields.")
                return

            duration = int(duration_str)
            sequence = int(sequence_str)
            budget = float(budget_str) if budget_str else 0.0

            section = ActivitySection.PRE_KICKOFF if section_str == ActivitySection.PRE_KICKOFF.value else ActivitySection.POST_KICKOFF

            # Create activity
            activity = Activity(
                task=task,
                action_needed=action,
                duration=duration,
                precursor=precursor,
                sequence=sequence,
                resources=resources,
                budget=budget,
                section=section
            )

            self.activities_data.append(activity)

            # Refresh the activities list to maintain proper sorting
            self.refresh_activities_list()

            # Clear form
            for var in self.form_vars.values():
                var.set("")

            messagebox.showinfo("Success", "Activity added successfully!")

        except ValueError as e:
            messagebox.showerror("Error", "Please enter valid numbers for Duration, Sequence, and Budget.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def setup_context_menu(self):
        """Setup right-click context menu for activities tree"""
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Edit Activity", command=self.edit_activity_with_dialog)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Delete Activity", command=self.delete_activity_with_confirmation)
        
        # Bind right-click event
        self.activities_tree.bind("<Button-3>", self.show_context_menu)  # Right-click
        
    def show_context_menu(self, event):
        """Show context menu on right-click"""
        # Select the item under cursor
        item = self.activities_tree.identify_row(event.y)
        if item:
            self.activities_tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)
    
    def delete_activity_with_confirmation(self):
        """Delete the selected activity with confirmation dialog"""
        selected_item = self.activities_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select an activity to delete.")
            return

        # Get activity details for confirmation
        item_values = self.activities_tree.item(selected_item[0])['values']
        activity_name = item_values[1] if len(item_values) > 1 else "this activity"
        
        # Show confirmation dialog
        confirm = messagebox.askyesno(
            "Confirm Deletion", 
            f"Are you sure you want to delete the activity:\n\n'{activity_name}'?\n\nThis action cannot be undone.",
            icon='warning'
        )
        
        if not confirm:
            return

        # Get the index of the selected item
        index = self.activities_tree.index(selected_item[0])

        # Remove from data and treeview
        if 0 <= index < len(self.activities_data):
            del self.activities_data[index]
            self.activities_tree.delete(selected_item[0])
            self.update_activities_count()
            messagebox.showinfo("Success", "Activity deleted successfully!")

    def edit_activity_with_dialog(self):
        """Edit the selected activity with a dialog window"""
        selected_item = self.activities_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select an activity to edit.")
            return

        # Get the index of the selected item
        index = self.activities_tree.index(selected_item[0])
        
        if not (0 <= index < len(self.activities_data)):
            messagebox.showerror("Error", "Invalid activity selection.")
            return
            
        # Get the activity to edit
        activity = self.activities_data[index]
        
        # Create edit dialog
        self.create_edit_dialog(activity, index)

    def create_edit_dialog(self, activity, index):
        """Create and show the edit dialog for an activity"""
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Activity")
        edit_window.geometry("600x500")
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # Center the dialog
        edit_window.geometry("+%d+%d" % (
            self.root.winfo_rootx() + 100,
            self.root.winfo_rooty() + 100
        ))
        
        # Main frame
        main_frame = ttk.Frame(edit_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Form fields
        fields = [
            ("Task:", "task_var"),
            ("Action Needed:", "action_var"),
            ("Duration (days):", "duration_var"),
            ("Precursor:", "precursor_var"),
            ("Sequence:", "sequence_var"),
            ("Resources:", "resources_var"),
            ("Budget:", "budget_var")
        ]

        edit_vars = {}
        for i, (label, var_name) in enumerate(fields):
            ttk.Label(main_frame, text=label).grid(row=i, column=0, sticky=tk.W, padx=(0, 10), pady=5)
            
            edit_vars[var_name] = tk.StringVar()
            entry = ttk.Entry(main_frame, textvariable=edit_vars[var_name], width=40)
            entry.grid(row=i, column=1, sticky=(tk.W, tk.E), pady=5)
            
        # Section selection
        ttk.Label(main_frame, text="Section:").grid(row=len(fields), column=0, sticky=tk.W, padx=(0, 10), pady=5)
        section_var = tk.StringVar()
        section_combo = ttk.Combobox(main_frame, textvariable=section_var,
                                   values=[s.value for s in ActivitySection], state="readonly", width=37)
        section_combo.grid(row=len(fields), column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Pre-fill the form with current activity data
        edit_vars['task_var'].set(activity.task)
        edit_vars['action_var'].set(activity.action_needed)
        edit_vars['duration_var'].set(str(activity.duration))
        edit_vars['precursor_var'].set(activity.precursor)
        edit_vars['sequence_var'].set(str(activity.sequence))
        edit_vars['resources_var'].set(activity.resources)
        edit_vars['budget_var'].set(str(activity.budget))
        section_var.set(activity.section.value)
        
        # Configure column weight
        main_frame.columnconfigure(1, weight=1)
        
        # Buttons frame
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=len(fields)+1, column=0, columnspan=2, pady=(20, 0))
        
        def save_changes():
            """Save the edited activity"""
            try:
                # Validate inputs
                task = edit_vars['task_var'].get().strip()
                action = edit_vars['action_var'].get().strip()
                duration_str = edit_vars['duration_var'].get().strip()
                precursor = edit_vars['precursor_var'].get().strip()
                sequence_str = edit_vars['sequence_var'].get().strip()
                resources = edit_vars['resources_var'].get().strip()
                budget_str = edit_vars['budget_var'].get().strip()
                section_str = section_var.get()

                if not all([task, action, duration_str, sequence_str]):
                    messagebox.showerror("Error", "Please fill in Task, Action Needed, Duration, and Sequence fields.")
                    return

                duration = int(duration_str)
                sequence = int(sequence_str)
                budget = float(budget_str) if budget_str else 0.0

                section = ActivitySection.PRE_KICKOFF if section_str == ActivitySection.PRE_KICKOFF.value else ActivitySection.POST_KICKOFF

                # Update the activity
                self.activities_data[index].task = task
                self.activities_data[index].action_needed = action
                self.activities_data[index].duration = duration
                self.activities_data[index].precursor = precursor
                self.activities_data[index].sequence = sequence
                self.activities_data[index].resources = resources
                self.activities_data[index].budget = budget
                self.activities_data[index].section = section

                # Update the treeview
                self.refresh_activities_list()

                edit_window.destroy()
                messagebox.showinfo("Success", "Activity updated successfully!")

            except ValueError:
                messagebox.showerror("Error", "Please enter valid numbers for Duration, Sequence, and Budget.")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {str(e)}")
        
        def cancel_edit():
            """Cancel editing and close dialog"""
            edit_window.destroy()
        
        ttk.Button(buttons_frame, text="Save Changes", command=save_changes).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(buttons_frame, text="Cancel", command=cancel_edit).pack(side=tk.LEFT)

    def clear_all_with_confirmation(self):
        """Clear all activities and reset the form with confirmation"""
        if len(self.activities_data) == 0:
            messagebox.showinfo("Information", "No data to clear.")
            return
            
        confirm = messagebox.askyesno(
            "Confirm Clear All", 
            f"Are you sure you want to clear all project data?\n\n"
            f"This will remove:\n"
            f"• Project title\n"
            f"• All {len(self.activities_data)} activities\n"
            f"• Form data\n\n"
            f"This action cannot be undone.",
            icon='warning'
        )
        
        if confirm:
            self.activities_data.clear()
            self.activities_tree.delete(*self.activities_tree.get_children())
            self.project_title_var.set("")
            for var in self.form_vars.values():
                var.set("")
            self.update_activities_count()
            messagebox.showinfo("Success", "All data cleared!")

    def generate_preview_text(self, project):
        """Generate preview text for both display and file output"""
        schedules = ScheduleCalculator.calculate_schedules(project)
        max_activities = ScheduleCalculator.get_max_duration_activities(project)

        preview_text = f"Project Schedule Summary\n"
        preview_text += f"=" * 50 + "\n\n"
        preview_text += f"Project Title: {project.title}\n"
        preview_text += f"Calendar Format: {project.calendar_format.value}\n"
        preview_text += f"Total Activities: {len(project.activities)}\n"
        
        # Budget summary
        total_budget = sum(activity.budget for activity in project.activities)
        preview_text += f"Total Budget: ${total_budget:,.2f}\n\n"

        for section in [ActivitySection.PRE_KICKOFF, ActivitySection.POST_KICKOFF]:
            activities = project.get_activities_by_section(section)
            if not activities:
                continue
                
            preview_text += f"{section.value} ({len(activities)} activities):\n"
            preview_text += f"-" * 40 + "\n"
            activities.sort(key=lambda x: x.sequence)

            current_sequence = None
            for activity in activities:
                if current_sequence != activity.sequence:
                    current_sequence = activity.sequence
                    raw_schedule = schedules.get(activity.sequence, 0)
                    adjusted_schedule = ScheduleCalculator.apply_calendar_format(raw_schedule, project.calendar_format)
                    preview_text += f"\nSequence {activity.sequence} (Schedule: {adjusted_schedule} days):\n"

                is_max = any(activity is max_act for max_act, _ in max_activities)
                max_indicator = " [MAX DURATION]" if is_max else ""
                
                preview_text += f"  • {activity.task}\n"
                preview_text += f"    Duration: {activity.duration} days{max_indicator}\n"
                preview_text += f"    Action: {activity.action_needed}\n"
                preview_text += f"    Resources: {activity.resources}\n"
                preview_text += f"    Budget: ${activity.budget:,.2f}\n"
                if activity.precursor:
                    preview_text += f"    Precursor: {activity.precursor}\n"
                preview_text += "\n"

        return preview_text

    def preview_schedule(self):
        """Show a preview of the schedule calculations"""
        if not self.validate_project():
            return

        project = self.create_project()
        preview_text = self.generate_preview_text(project)

        # Show preview window
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Schedule Preview")
        preview_window.geometry("700x500")
        
        # Create main frame
        main_frame = ttk.Frame(preview_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create text frame with scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, padx=10, pady=10)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        text_widget.insert("1.0", preview_text)
        text_widget.config(state=tk.DISABLED)

        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.config(yscrollcommand=scrollbar.set)
        
        # Create buttons frame
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, pady=(10, 0))
        
        def copy_to_clipboard():
            """Copy the preview text to clipboard"""
            try:
                preview_window.clipboard_clear()
                preview_window.clipboard_append(preview_text)
                preview_window.update()  # Required on some systems
                messagebox.showinfo("Success", "Schedule preview copied to clipboard!", parent=preview_window)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to copy to clipboard: {str(e)}", parent=preview_window)
        
        def close_preview():
            """Close the preview window"""
            preview_window.destroy()
        
        # Add buttons
        ttk.Button(buttons_frame, text="Copy to Clipboard", command=copy_to_clipboard).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(buttons_frame, text="Close", command=close_preview).pack(side=tk.LEFT)

    def generate_excel(self):
        """Generate and save the Excel file"""
        if not self.validate_project():
            return

        try:
            # Create default filename based on project title
            project_title = self.project_title_var.get().strip()
            safe_title = "".join(c for c in project_title if c.isalnum() or c in (' ', '-', '_')).strip()
            default_filename = f"{safe_title}_schedule.xlsx" if safe_title else "project_schedule.xlsx"

            # Get output file path with improved dialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Project Schedule As...",
                initialfile=default_filename
            )

            if not file_path:
                return  # User cancelled

            # Show progress message
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Generating Excel...")
            progress_window.geometry("300x100")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Center the progress window
            progress_window.geometry("+%d+%d" % (
                self.root.winfo_rootx() + 50,
                self.root.winfo_rooty() + 50
            ))
            
            ttk.Label(progress_window, text="Generating Excel file...", font=('Arial', 12)).pack(expand=True)
            progress_window.update()

                         # Generate Excel file
            project = self.create_project()
            generator = ExcelGenerator(project)
            generator.generate(file_path)
            
            # Generate accompanying .txt file with schedule details
            txt_file_path = file_path.replace('.xlsx', '_details.txt')
            try:
                preview_text = self.generate_preview_text(project)
                with open(txt_file_path, 'w', encoding='utf-8') as txt_file:
                    txt_file.write(preview_text)
            except Exception as txt_error:
                print(f"Warning: Could not create .txt file: {txt_error}")
            
            # Close progress window
            progress_window.destroy()

            # Show success message with option to open file location
            files_created = f"Files created:\n• {file_path}\n• {txt_file_path}"
            result = messagebox.askyesno(
                "Success", 
                f"Project files generated successfully!\n\n"
                f"{files_created}\n\n"
                f"Would you like to open the file location?",
                icon='info'
            )
            
            if result:
                import os
                import subprocess
                import platform
                
                # Open file location in file explorer
                if platform.system() == "Windows":
                    subprocess.run(['explorer', '/select,', file_path], shell=False)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(["open", "-R", file_path])
                else:  # Linux
                    subprocess.run(["xdg-open", os.path.dirname(file_path)])

        except Exception as e:
            # Close progress window if it exists
            try:
                progress_window.destroy()
            except:
                pass
            messagebox.showerror("Error", f"Failed to generate Excel file:\n{str(e)}")

    def validate_project(self):
        """Validate project data before processing"""
        if not self.project_title_var.get().strip():
            messagebox.showerror("Error", "Please enter a project title.")
            return False

        if not self.activities_data:
            messagebox.showerror("Error", "Please add at least one activity.")
            return False

        return True

    def create_project(self):
        """Create a Project object from the GUI data"""
        # Get calendar format
        calendar_format = CalendarFormat.FIVE_DAY
        for fmt in CalendarFormat:
            if fmt.value == self.calendar_format_var.get():
                calendar_format = fmt
                break

        # Create project
        project = Project(
            title=self.project_title_var.get().strip(),
            calendar_format=calendar_format
        )

        # Add activities
        for activity in self.activities_data:
            project.add_activity(activity)

        return project
    
    def update_activities_count(self):
        """Update the activities list frame title with current count"""
        count = len(self.activities_data)
        self.list_frame.config(text=f"Activities List ({count} activities)")

    def refresh_activities_list(self):
        """Refresh the activities list with current data, sorted by section and sequence"""
        # Clear existing items
        self.activities_tree.delete(*self.activities_tree.get_children())
        
        # Sort activities: first by section (Pre-Kickoff first), then by sequence
        sorted_activities = sorted(self.activities_data, key=lambda x: (
            0 if x.section == ActivitySection.PRE_KICKOFF else 1,  # Pre-Kickoff first
            x.sequence
        ))
        
        # Repopulate the treeview
        for activity in sorted_activities:
            budget_millions = activity.budget / 1000000 if activity.budget > 0 else 0.0
            self.activities_tree.insert("", tk.END, values=(
                activity.section.value, 
                activity.task, 
                activity.action_needed, 
                activity.duration, 
                activity.precursor, 
                activity.sequence, 
                activity.resources, 
                f"{budget_millions:.2f}"
            ))
        
        # Update the count
        self.update_activities_count()

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()

    def load_excel(self):
        """Load an existing Excel file generated by the scheduler"""
        file_path = filedialog.askopenfilename(
            title="Load Project Excel File",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ],
            initialdir=os.getcwd()
        )
        
        if not file_path:
            return
        
        try:
            # Show loading progress
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Loading Excel File")
            progress_window.geometry("300x100")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Center the progress window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (300 // 2)
            y = (progress_window.winfo_screenheight() // 2) - (100 // 2)
            progress_window.geometry(f"300x100+{x}+{y}")
            
            progress_label = ttk.Label(progress_window, text="Loading Excel file...")
            progress_label.pack(expand=True)
            
            progress_window.update()
            
            # Load the Excel file
            loader = ExcelLoader(file_path)
            project = loader.load_project()
            
            # Clear existing data
            self.clear_all_data()
            
            # Load project data into GUI
            self.project_title_var.set(project.title)
            self.calendar_format_var.set(project.calendar_format.value)
            
            # Load activities
            for activity in project.activities:
                self.activities_data.append(activity)
                
                # Add to treeview (show budget in millions)
                budget_millions = activity.budget / 1000000 if activity.budget > 0 else 0.0
                self.activities_tree.insert("", tk.END, values=(
                    activity.section.value, 
                    activity.task, 
                    activity.action_needed, 
                    activity.duration, 
                    activity.precursor, 
                    activity.sequence, 
                    activity.resources, 
                    f"{budget_millions:.2f}"
                ))
            
            # Update activities count
            self.update_activities_count()
            
            # Close progress window
            progress_window.destroy()
            
            # Show success message
            messagebox.showinfo(
                "Load Successful", 
                f"Successfully loaded project:\n\n"
                f"Title: {project.title}\n"
                f"Activities: {len(project.activities)}\n"
                f"Calendar Format: {project.calendar_format.value}\n\n"
                f"You can now edit activities or add new ones!"
            )
            
        except Exception as e:
            # Close progress window if still open
            try:
                progress_window.destroy()
            except:
                pass
            
            messagebox.showerror(
                "Load Error", 
                f"Failed to load Excel file:\n\n{str(e)}\n\n"
                f"Please ensure the file was generated by this scheduler program."
            )
    
    def clear_all_data(self):
        """Clear all data from the GUI without confirmation"""
        # Clear form fields
        self.project_title_var.set("")
        
        # Clear form variables using the form_vars dictionary
        for var in self.form_vars.values():
            var.set("")
        
        # Clear section dropdown
        self.section_var.set("Pre-Kickoff Activities")
        
        # Clear activities data and treeview
        self.activities_data.clear()
        for item in self.activities_tree.get_children():
            self.activities_tree.delete(item)
        
        # Update count
        self.update_activities_count()


def main():
    """Main function to run the Project Scheduler application"""
    app = ProjectSchedulerGUI()
    app.run()


if __name__ == "__main__":
    main() 