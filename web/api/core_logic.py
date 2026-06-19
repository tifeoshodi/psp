"""
Project Scheduler Application
Creates Excel schedules from user input with advanced formatting and calculations.
"""

from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from enum import Enum
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import subprocess
import platform
from dataclasses import dataclass
from enum import Enum
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import os
import sys
import datetime
from typing import Union


def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class CalendarFormat(Enum):
    FIVE_DAY = "5-day week"
    SIX_DAY = "6-day week"
    SEVEN_DAY = "7-day week"


class ActivitySection(Enum):
    PRE_KICKOFF = "Pre-Kickoff Activities"
    POST_KICKOFF = "Post Kick-off Activities"


@dataclass
class Activity:
    """Represents a single project activity/task"""
    task: str
    action_needed: str
    duration: int
    precursor: str
    sequence: int
    resources: str
    budget: float
    section: ActivitySection

    def __post_init__(self):
        # Ensure duration is positive
        if self.duration < 0:
            self.duration = 0


@dataclass
class Project:
    """Represents the entire project with all activities"""
    title: str
    activities: List[Activity] = field(default_factory=list)
    calendar_format: CalendarFormat = CalendarFormat.FIVE_DAY
    start_date: Optional[datetime.date] = None

    def add_activity(self, activity: Activity):
        """Add an activity to the project"""
        self.activities.append(activity)

    def get_activities_by_section(self, section: ActivitySection) -> List[Activity]:
        """Get all activities in a specific section"""
        return [activity for activity in self.activities if activity.section == section]

    def get_sequences_by_section(self, section: ActivitySection) -> List[int]:
        """Get unique sequences in a section, sorted"""
        activities = self.get_activities_by_section(section)
        sequences = list(set(activity.sequence for activity in activities))
        return sorted(sequences)


class ScheduleCalculator:
    """Handles schedule calculations and calendar adjustments"""

    @staticmethod
    def calculate_schedules(project: Project) -> Dict[int, int]:
        """Calculate schedule values for each sequence"""
        schedules = {}

        # Pre-kickoff activities always have schedule = 0
        pre_kickoff_sequences = project.get_sequences_by_section(ActivitySection.PRE_KICKOFF)
        for seq in pre_kickoff_sequences:
            schedules[seq] = 0

        # Post-kickoff activities have incremental schedules
        post_kickoff_sequences = project.get_sequences_by_section(ActivitySection.POST_KICKOFF)
        post_kickoff_activities = project.get_activities_by_section(ActivitySection.POST_KICKOFF)

        cumulative_schedule = 0
        for seq in post_kickoff_sequences:
            # Find max duration in this sequence
            seq_activities = [a for a in post_kickoff_activities if a.sequence == seq]
            max_duration = max(a.duration for a in seq_activities) if seq_activities else 0

            # Add max duration to cumulative schedule
            cumulative_schedule += max_duration
            schedules[seq] = cumulative_schedule

        return schedules

    @staticmethod
    def apply_calendar_format(schedule_days: int, calendar_format: CalendarFormat) -> int:
        """Apply calendar format adjustments to schedule days"""
        if calendar_format == CalendarFormat.FIVE_DAY:
            # Add 2 days per week for weekends
            weeks = schedule_days // 5
            extra_days = weeks * 2
            return schedule_days + extra_days
        elif calendar_format == CalendarFormat.SIX_DAY:
            # Add 1 day per week for single non-working day
            weeks = schedule_days // 6
            extra_days = weeks * 1
            return schedule_days + extra_days
        else:  # SEVEN_DAY
            return schedule_days

    @staticmethod
    def get_max_duration_activities(project: Project) -> List[Tuple[Activity, int]]:
        """Get activities that have max duration within their sequence group"""
        max_duration_activities = []

        for section in [ActivitySection.PRE_KICKOFF, ActivitySection.POST_KICKOFF]:
            sequences = project.get_sequences_by_section(section)
            section_activities = project.get_activities_by_section(section)

            for seq in sequences:
                seq_activities = [a for a in section_activities if a.sequence == seq]
                if seq_activities:
                    max_duration = max(a.duration for a in seq_activities)
                    for activity in seq_activities:
                        if activity.duration == max_duration:
                            max_duration_activities.append((activity, max_duration))

        return max_duration_activities


class ExcelGenerator:
    """Generates Excel files with proper formatting and calculations"""

    def __init__(self, project: Project, custom_logo_path: Optional[str] = None):
        self.project = project
        self.custom_logo_path = custom_logo_path
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Project Schedule"

        # Styling constants
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True)
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self, output_path: str):
        """Generate the complete Excel file"""
        current_row = 1

        # Add formatted header (project title, logo space, and timestamp)
        current_row = self._add_formatted_header(current_row)
        # No extra row - headers should be consecutive

        # Add column headers
        current_row = self._add_styled_headers(current_row)

        # Add Pre-Kickoff Activities section
        current_row = self._add_section(ActivitySection.PRE_KICKOFF, current_row)
        current_row = self._add_empty_row(current_row)  # Merged empty row

        # Add Post Kick-off Activities section
        current_row = self._add_section(ActivitySection.POST_KICKOFF, current_row)
        
        # Add budget total row
        current_row = self._add_budget_total(current_row)

        # Apply formatting
        self._apply_formatting()

        # Apply sheet protection (allow editing only for Review Comments column)
        self._apply_sheet_protection()

        # Remove gridlines outside content area
        self._remove_external_gridlines()

        # Apply freeze panes for fixed headers
        self._apply_freeze_panes()

        # Generate charts - DISABLED FOR NOW (to be revisited)
        # self._generate_charts(current_row)

        # Generate Gantt chart as second worksheet
        gantt_generator = GanttChartGenerator(self.project, self.workbook)
        gantt_generator.generate_gantt_chart()

        # Save file
        self.workbook.save(output_path)

    def _add_formatted_header(self, start_row: int) -> int:
        """Add formatted header with logo space, project title, and timestamp"""
        import datetime
        import os
        
        # Merge rows 1 & 2 for header
        # Merge cells A-B for logo space and timestamp
        self.worksheet.merge_cells(f"A{start_row}:B{start_row + 1}")
        
        # Add logo and timestamp to A-B cell
        self._add_logo_and_timestamp(start_row)
        
        # Merge cells C-I for project title only (no timestamp)
        self.worksheet.merge_cells(f"C{start_row}:I{start_row + 1}")
        
        # Add project title only
        title_cell = self.worksheet.cell(row=start_row, column=3, value=self.project.title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Set background color for A-B merged cell
        logo_cell = self.worksheet.cell(row=start_row, column=1)
        logo_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Merge cell J for empty column (no borders)
        self.worksheet.merge_cells(f"J{start_row}:J{start_row + 1}")
        empty_cell = self.worksheet.cell(row=start_row, column=10, value="")
        empty_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        empty_cell.border = None  # Remove border
        
        # Merge cell K for Review Comments space (no borders)
        self.worksheet.merge_cells(f"K{start_row}:K{start_row + 1}")
        review_space_cell = self.worksheet.cell(row=start_row, column=11, value="")
        review_space_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        review_space_cell.border = None  # Remove border
        
        # Set row heights for header rows (halved from doubled)
        self.worksheet.row_dimensions[start_row].height = 35
        self.worksheet.row_dimensions[start_row + 1].height = 35
        
        return start_row + 2  # Skip both merged rows

    def _add_logo_and_timestamp(self, start_row: int):
        """Add logo and timestamp to the merged A-B cell"""
        import os
        import datetime
        import getpass
        
        # Get user account name
        try:
            user_account = getpass.getuser()
        except:
            user_account = os.environ.get('USERNAME', os.environ.get('USER', 'Unknown User'))
        
        # Add timestamp text to the merged cell with new format
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        timestamp_text = f"Generated by: {user_account}\n{timestamp}"
        
        timestamp_cell = self.worksheet.cell(row=start_row, column=1, value=timestamp_text)
        timestamp_cell.font = Font(size=9, bold=False)
        timestamp_cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        # Use custom logo if provided, otherwise use default
        if self.custom_logo_path and os.path.exists(self.custom_logo_path):
            logo_path = self.custom_logo_path
            print(f"Using custom logo: {logo_path}")
        else:
            # Look for default logo file - handle both development and executable paths
            logo_path = get_resource_path("IESL-Logo.png")
            print(f"Using default logo: {logo_path}")
        
        if os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image
                
                # Try to use PIL for better image handling
                try:
                    from PIL import Image as PILImage
                    
                    # Load the image to get original dimensions
                    pil_img = PILImage.open(logo_path)
                    original_width, original_height = pil_img.size
                    
                    # Calculate target dimensions while maintaining aspect ratio
                    # Target area: approximately 120 pixels wide, 60 pixels high (for 2-row merge)
                    target_max_width = 120
                    target_max_height = 60
                    
                    # Calculate scaling factor to maintain aspect ratio
                    width_ratio = target_max_width / original_width
                    height_ratio = target_max_height / original_height
                    scale_factor = min(width_ratio, height_ratio)
                    
                    # Calculate new dimensions
                    new_width = int(original_width * scale_factor)
                    new_height = int(original_height * scale_factor)
                    
                    print(f"Logo resized from {original_width}x{original_height} to {new_width}x{new_height}")
                    
                except ImportError:
                    # Fallback to reasonable default sizes if PIL is not available
                    print("PIL not available, using default logo sizing")
                    new_width = 100  # Larger than original fixed size
                    new_height = 50  # Better proportions
                
                # Load and resize the image for Excel
                img = Image(logo_path)
                img.width = new_width
                img.height = new_height
                
                # Position the image in cell A1 (top part of the merged cell)
                img.anchor = f"A{start_row}"
                
                # Add image to worksheet
                self.worksheet.add_image(img)
                
            except Exception as e:
                print(f"Warning: Could not add logo image: {e}")
        else:
            print(f"Warning: Logo file not found at: {logo_path}")
            # Only try fallback if using default logo (not custom)
            if not self.custom_logo_path:
                # Also check if file exists in current directory as fallback
                fallback_path = "IESL-Logo.png"
                if os.path.exists(fallback_path):
                    print(f"Found logo in current directory, using: {fallback_path}")
                    try:
                        from openpyxl.drawing.image import Image
                        img = Image(fallback_path)
                        img.width = 100  # Default size
                        img.height = 50
                        img.anchor = f"A{start_row}"
                        self.worksheet.add_image(img)
                    except Exception as e:
                        print(f"Warning: Could not add fallback logo: {e}")
            else:
                print("Custom logo file not found - proceeding without logo")

    def _add_styled_headers(self, start_row: int) -> int:
        """Add styled column headers"""
        headers = [
            "S/No", "Activities/Tasks", "Action Needed", "Duration (in days)", "Precursor",
            "Sequence", "Schedule (in days)", "Resources", "Budget (MILLION)", "", "Review Comments"
        ]
        
        # Light Yellow background for columns A-I (Custom Light Yellow 13)
        light_yellow_fill = PatternFill(start_color="FFFDD0", end_color="FFFDD0", fill_type="solid")
        # White background for empty column J
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # Red background for column K (Review Comments)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=start_row, column=col, value=header)
            cell.font = Font(size=11, bold=True, color="000000" if col <= 10 else "FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Apply borders to all columns except empty column J and Review Comments column K
            if col != 10 and col != 11:
                cell.border = self.border
            
            # Apply appropriate background color
            if col <= 9:  # Columns A-I
                cell.fill = light_yellow_fill
            elif col == 10:  # Empty column J
                cell.fill = white_fill
                # No border applied for empty column
            else:  # Column K (Review Comments)
                cell.fill = red_fill
        
        # Set header row height (halved from doubled)
        self.worksheet.row_dimensions[start_row].height = 25
        
        return start_row + 1

    def _add_section(self, section: ActivitySection, start_row: int) -> int:
        """Add a section with its activities"""
        # Determine background color based on section type
        if section == ActivitySection.PRE_KICKOFF:
            section_color = "ffff99"  # Yellow for Pre-kickoff Activities
        else:  # POST_KICKOFF
            section_color = "D2E3A3"  # Olive Green, Accent 3, Lighter 80% for Post kick-off Activities
        
        # Add section header with appropriate background (left-formatted, excluding Review Comments column)
        section_cell = self.worksheet.cell(row=start_row, column=1, value=section.value)
        section_cell.fill = PatternFill(start_color=section_color, end_color=section_color, fill_type="solid")
        section_cell.font = Font(size=12, bold=True)
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        section_cell.border = self.border
        self.worksheet.merge_cells(f"A{start_row}:I{start_row}")  # Merge only to column I, not affecting Review Comments
        
        # Keep column J empty with no border
        j_cell = self.worksheet.cell(row=start_row, column=10, value="")
        j_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        j_cell.border = None
        
        # Keep Review Comments column K separate with no border
        k_cell = self.worksheet.cell(row=start_row, column=11, value="")
        k_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        k_cell.border = None
        
        # Set section header row height (halved from doubled)
        self.worksheet.row_dimensions[start_row].height = 22
        
        current_row = start_row + 1

        # Get activities and schedules for this section
        activities = self.project.get_activities_by_section(section)
        schedules = ScheduleCalculator.calculate_schedules(self.project)
        max_duration_activities = ScheduleCalculator.get_max_duration_activities(self.project)
        max_duration_set = {id(activity) for activity, _ in max_duration_activities}

        # Sort activities by sequence
        activities.sort(key=lambda x: x.sequence)

        # Add activities with S/N numbering
        activity_number = 1
        for activity in activities:
            self._add_activity_row(activity, current_row, schedules, max_duration_set, activity_number)
            current_row += 1
            activity_number += 1

        # Merge cells for same sequences
        self._merge_schedule_cells(section, start_row + 1, current_row - 1)

        return current_row

    def _add_empty_row(self, start_row: int) -> int:
        """Add an empty row with merged columns (excluding Review Comments column)"""
        # Merge columns A to I only (not affecting Review Comments column K)
        self.worksheet.merge_cells(f"A{start_row}:I{start_row}")
        empty_cell = self.worksheet.cell(row=start_row, column=1, value="")
        empty_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Keep column J empty with no border
        j_cell = self.worksheet.cell(row=start_row, column=10, value="")
        j_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        j_cell.border = None
        
        # Keep Review Comments column K separate with no border
        k_cell = self.worksheet.cell(row=start_row, column=11, value="")
        k_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        k_cell.border = None
        
        # Set row height for empty row (back to reasonable size)
        self.worksheet.row_dimensions[start_row].height = 15
        
        return start_row + 1

    def _add_activity_row(self, activity: Activity, row: int, schedules: Dict[int, int], max_duration_set: set, activity_number: int):
        """Add a single activity row with enhanced formatting"""
        # Calculate adjusted schedule - Pre-Kickoff activities always have 0
        if activity.section == ActivitySection.PRE_KICKOFF:
            adjusted_schedule = 0
            schedule_formula = "0"  # Simple value for pre-kickoff
        else:
            raw_schedule = schedules.get(activity.sequence, 0)
            adjusted_schedule = ScheduleCalculator.apply_calendar_format(raw_schedule, self.project.calendar_format)
            
            # Create formula comment for schedule calculation
            if self.project.calendar_format == CalendarFormat.FIVE_DAY:
                schedule_formula = f"={raw_schedule}+({raw_schedule}//5)*2"  # Add weekends
            elif self.project.calendar_format == CalendarFormat.SIX_DAY:
                schedule_formula = f"={raw_schedule}+({raw_schedule}//6)*1"  # Add one day per week
            else:
                schedule_formula = f"={raw_schedule}"  # No adjustment for 7-day

        # Format budget to millions with 2 decimal places
        budget_millions = activity.budget / 1000000 if activity.budget > 0 else 0.0

        values = [
            activity_number,  # S/N column
            activity.task,
            activity.action_needed,
            activity.duration,
            activity.precursor,
            activity.sequence,
            adjusted_schedule,  # We'll add formula separately
            activity.resources,
            budget_millions,
            "",  # Empty column J
            ""   # Empty review comments column K
        ]

        # Determine if this row needs extra height (text-heavy columns) - increase all heights
        text_heavy_columns = [activity.task, activity.action_needed, activity.precursor]
        max_text_length = max(len(str(text)) for text in text_heavy_columns)
        
        # Set row height based on content length (halved from doubled)
        if max_text_length > 50:
            self.worksheet.row_dimensions[row].height = 40
        elif max_text_length > 30:
            self.worksheet.row_dimensions[row].height = 35
        else:
            self.worksheet.row_dimensions[row].height = 30

        # Determine background color for S/No column based on section
        if activity.section == ActivitySection.PRE_KICKOFF:
            sno_bg_color = "ffff99"  # Yellow for Pre-kickoff Activities
        else:  # POST_KICKOFF
            sno_bg_color = "D2E3A3"  # Olive Green, Accent 3, Lighter 80% for Post kick-off Activities

        for col, value in enumerate(values, 1):
            cell = self.worksheet.cell(row=row, column=col, value=value)
            
            # Apply borders consistently (except for empty column J)
            if col != 10:  # All columns except empty column J
                cell.border = self.border
            
            # Apply specific formatting based on column
            if col == 1:  # S/N column with section-specific background color
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color=sno_bg_color, end_color=sno_bg_color, fill_type="solid")
            elif col in [2, 3]:  # Activities/Tasks, Action Needed columns
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.font = Font(size=10)
            elif col in [4, 6, 7]:  # Duration, Sequence, Schedule columns
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)
            elif col in [5, 8]:  # Precursor, Resources columns (center-formatted as requested)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(size=10)
            elif col == 9:  # Budget column  
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '"$"#,##0.00'  # Currency format with dollar sign and thousands separators
                cell.font = Font(size=10)
            elif col == 10:  # Empty column J
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.border = None  # No border for empty column
            elif col == 11:  # Review Comments column K
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.font = Font(size=10)
                # Unlock this cell for editing
                from openpyxl.styles import Protection
                cell.protection = Protection(locked=False)

            # Highlight max duration text in red
            if col == 4 and id(activity) in max_duration_set:  # Duration column (now col 4)
                cell.font = Font(color="FF0000", bold=True, size=10)  # Red text, bold
        
        # Add formula to schedule cell (column 7) if it's not pre-kickoff
        if activity.section != ActivitySection.PRE_KICKOFF:
            schedule_cell = self.worksheet.cell(row=row, column=7)
            # Add comment with formula explanation
            from openpyxl.comments import Comment
            comment_text = f"Formula: {schedule_formula}\nBase schedule: {schedules.get(activity.sequence, 0)} days\nCalendar format: {self.project.calendar_format.value}"
            comment = Comment(comment_text, "Project Scheduler")
            schedule_cell.comment = comment

    def _merge_schedule_cells(self, section: ActivitySection, start_row: int, end_row: int):
        """Merge cells in Schedule column for same sequence values"""
        activities = self.project.get_activities_by_section(section)
        sequences = self.project.get_sequences_by_section(section)

        for seq in sequences:
            seq_activities = [a for a in activities if a.sequence == seq]
            if len(seq_activities) > 1:
                # Find the rows for this sequence
                seq_rows = []
                current_activities = activities[:]
                current_activities.sort(key=lambda x: x.sequence)

                row_counter = start_row
                for activity in current_activities:
                    if activity.sequence == seq:
                        seq_rows.append(row_counter)
                    row_counter += 1

                if len(seq_rows) > 1:
                    first_row = min(seq_rows)
                    last_row = max(seq_rows)
                    self.worksheet.merge_cells(f"G{first_row}:G{last_row}")  # Schedule is still column G

    def _add_budget_total(self, start_row: int) -> int:
        """Add a total row for budget calculation"""
        current_row = start_row + 1  # Add empty row before total
        
        # Add "Total" label (now in Resources column)
        total_cell = self.worksheet.cell(row=current_row, column=8, value="Total:")
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_cell.border = self.border
        
        # Calculate total budget in millions
        total_budget_millions = sum(activity.budget for activity in self.project.activities) / 1000000
        budget_cell = self.worksheet.cell(row=current_row, column=9, value=total_budget_millions)
        budget_cell.font = Font(bold=True)
        budget_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        budget_cell.alignment = Alignment(horizontal='center', vertical='center')
        budget_cell.number_format = '"$"#,##0.00'  # Currency format with dollar sign and thousands separators
        budget_cell.border = self.border
        
        return current_row + 1

    def _apply_formatting(self):
        """Apply general formatting to the worksheet"""
        # Set specific column widths for better presentation (halved from doubled)
        column_widths = {
            'A': 8,   # S/No - keep thin (halved)
            'B': 30,  # Activities/Tasks - wider for text (halved)
            'C': 30,  # Action Needed - wider for text (halved)
            'D': 19,  # Duration (in days) - moderate (halved)
            'E': 25,  # Precursor - wider for text (halved)
            'F': 12,  # Sequence - narrow (halved)
            'G': 18,  # Schedule - moderate (halved)
            'H': 18,  # Resources - moderate (halved)
            'I': 18,  # Budget - moderate (halved)
            'J': 3,   # Empty column - very narrow (halved)
            'K': 25,  # Review Comments - wider for text (halved)
        }
        
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width

        # Apply borders consistently to all content cells
        max_row = self.worksheet.max_row
        max_col = 11  # K column (Review Comments)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                # Skip borders for empty column J only
                if col == 10:  # Empty column J - never has borders
                    continue
                # Apply borders to all other cells (including Review Comments column)
                if cell.value is not None or (row >= 1 and col <= max_col and col != 10):
                    if not hasattr(cell, 'border') or cell.border is None:  # Only apply if not already set
                        cell.border = self.border

    def _apply_sheet_protection(self):
        """Apply sheet protection, leaving only Review Comments column unlocked"""
        from openpyxl.styles import Protection
        
        # Lock all cells by default
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.protection = Protection(locked=True)
        
        # Unlock Review Comments column (column K, which is column 11)
        for row in self.worksheet.iter_rows(min_col=11, max_col=11):
            for cell in row:
                # Only unlock data rows (skip headers and section headers)
                if cell.row > 3:  # Assuming headers start around row 3
                    cell.protection = Protection(locked=False)
        
        # Protect the worksheet with a password (optional)
        # You can change or remove the password as needed
        self.worksheet.protection.sheet = True
        self.worksheet.protection.enable()

    def _remove_external_gridlines(self):
        """Remove gridlines outside the content area to make it plain white"""
        # Hide gridlines for the entire worksheet
        self.worksheet.sheet_view.showGridLines = False

    def _apply_freeze_panes(self):
        """Apply freeze panes to keep headers fixed while scrolling"""
        # Freeze the first 3 rows (header rows 1-2 + column headers row 3)
        # This will keep them visible when scrolling
        self.worksheet.freeze_panes = 'A4'  # Freeze everything above row 4

    # CHART GENERATION METHODS - DISABLED FOR NOW (to be revisited)
    
    # def _generate_charts(self, start_row: int):
    #     """Generate Gantt chart and S-curve on the same sheet"""
    #     try:
    #         # Calculate chart data (no date dependency needed)
    #         chart_data = self._prepare_chart_data()
    #         
    #         # Create new worksheet for chart data tables
    #         self.chart_worksheet = self.workbook.create_sheet("Chart Data")
    #         
    #         # Add title to chart data worksheet
    #         title_cell = self.chart_worksheet.cell(row=1, column=1, value="Project Chart Data Tables")
    #         title_cell.font = Font(size=16, bold=True)
    #         self.chart_worksheet.merge_cells("A1:E1")
    #         
    #         # Set column widths for chart data worksheet
    #         chart_column_widths = {
    #             'A': 40,  # Activity names
    #             'B': 15,  # Start Time
    #             'C': 12,  # Duration
    #             'D': 25,  # Section
    #             'E': 15,  # Critical Path
    #         }
    #         
    #         for col_letter, width in chart_column_widths.items():
    #             self.chart_worksheet.column_dimensions[col_letter].width = width
    #         
    #         # Position charts below the main schedule table (no data tables on main sheet)
    #         chart_start_row = start_row + 2
    #         
    #         # Generate Gantt chart
    #         gantt_end_row = self._create_gantt_chart(chart_data, chart_start_row)
    #         
    #         # Generate S-curve below Gantt chart
    #         s_curve_start_row = gantt_end_row + 2
    #         self._create_s_curve(chart_data, s_curve_start_row)
    #         
    #     except Exception as e:
    #         print(f"Warning: Could not generate charts: {e}")

    # def _prepare_chart_data(self):
    #     """Prepare data needed for both Gantt chart and S-curve"""
    #     # Calculate schedules for all activities
    #     schedules = ScheduleCalculator.calculate_schedules(self.project)
    #     
    #     # Get all activities sorted by section and sequence
    #     all_activities = []
    #     for section in [ActivitySection.PRE_KICKOFF, ActivitySection.POST_KICKOFF]:
    #         activities = self.project.get_activities_by_section(section)
    #         activities.sort(key=lambda x: x.sequence)
    #         all_activities.extend(activities)
    #     
    #     # Calculate critical path using existing max duration logic
    #     max_duration_activities = ScheduleCalculator.get_max_duration_activities(self.project)
    #     critical_path_activities = {id(activity) for activity, _ in max_duration_activities}
    #     
    #     chart_data = {
    #         'activities': [],
    #         'progress_timeline': []
    #     }
    #     
    #     # Process each activity for Gantt chart
    #     for activity in all_activities:
    #         # Calculate schedule position (start time in project timeline)
    #         if activity.section == ActivitySection.PRE_KICKOFF:
    #             # Pre-kickoff activities have negative start times
    #             start_time = -schedules.get(activity.sequence, 0)
    #             end_time = start_time + activity.duration
    #         else:
    #             # Post-kickoff activities start at their scheduled time
    #             start_time = schedules.get(activity.sequence, 0)
    #             end_time = start_time + activity.duration
    #         
    #         # Check if activity is on critical path
    #         is_critical = id(activity) in critical_path_activities
    #         
    #         # Add to activities data
    #         chart_data['activities'].append({
    #             'name': activity.task,
    #             'start_time': start_time,
    #             'end_time': end_time,
    #             'duration': activity.duration,
    #             'budget': activity.budget,
    #             'section': activity.section.value,
    #             'sequence': activity.sequence,
    #             'is_critical': is_critical
    #         })
    #     
    #     # Create ideal progress timeline for S-curve
    #     # Sort activities by end time to create cumulative progress
    #     sorted_activities = sorted(chart_data['activities'], key=lambda x: x['end_time'])
    #     
    #     total_activities = len(sorted_activities)
    #     total_duration = sum(act['duration'] for act in sorted_activities)
    #     
    #     cumulative_progress = 0
    #     cumulative_duration = 0
    #     
    #     for i, activity in enumerate(sorted_activities):
    #         cumulative_progress = ((i + 1) / total_activities) * 100  # Activity completion percentage
    #         cumulative_duration += activity['duration']
    #         duration_progress = (cumulative_duration / total_duration) * 100 if total_duration > 0 else 0
    #         
    #         chart_data['progress_timeline'].append({
    #             'time': activity['end_time'],
    #             'activity_progress': cumulative_progress,
    #             'duration_progress': duration_progress,
    #             'cumulative_duration': cumulative_duration
    #         })
    #     
    #     return chart_data



    # def _create_gantt_chart(self, chart_data, start_row: int) -> int:
    #     """Create a Gantt chart showing all activities with critical path highlighted"""
    #     try:
    #         from openpyxl.chart import BarChart, Reference, Series
    #         
    #         # Add title on main sheet
    #         title_cell = self.worksheet.cell(row=start_row, column=1, value="Project Gantt Chart - All Activities")
    #         title_cell.font = Font(size=14, bold=True)
    #         
    #         # Create data table on chart worksheet
    #         self._create_gantt_data_table(chart_data)
    #         
    #         # Create chart on main sheet but reference data from chart worksheet
    #         chart = BarChart()
    #         chart.type = "bar"
    #         chart.style = 12
    #         chart.title = "Project Timeline - Critical Path (Red) vs Normal Activities (Grey)"
    #         chart.y_axis.title = "Activities"
    #         chart.x_axis.title = "Duration (Days)"
    #         
    #         # Reference data from chart worksheet (starting from row 5, after headers)
    #         duration_data = Reference(self.chart_worksheet, min_col=3, min_row=5,
    #                                 max_col=3, max_row=4 + len(chart_data['activities']))
    #         cats = Reference(self.chart_worksheet, min_col=1, min_row=5,
    #                        max_col=1, max_row=4 + len(chart_data['activities']))
    #         
    #         # Add duration data as bars
    #         chart.add_data(duration_data, titles_from_data=False)
    #         chart.set_categories(cats)
    #         
    #         # Style the chart with conditional colors
    #         if chart.series:
    #             s1 = chart.series[0]
    #             # Set default grey color for all bars
    #             s1.graphicalProperties.solidFill = "808080"  # Grey for normal activities
    #             
    #             # Note: Individual bar coloring for critical path (red) would require 
    #             # more complex chart manipulation or separate series
    #         
    #         # Position the chart
    #         chart.anchor = f"A{start_row + 2}"
    #         chart.width = 25
    #         chart.height = max(15, len(chart_data['activities']) * 0.4)  # Scale height with number of activities
    #         
    #         self.worksheet.add_chart(chart)
    #         
    #         return start_row + int(chart.height / 15) + 3  # Estimate end row based on chart height
    #         
    #     except Exception as e:
    #         print(f"Warning: Could not create Gantt chart: {e}")
    #         return start_row + 10

    # def _create_gantt_data_table(self, chart_data):
    #     """Create Gantt chart data table on the chart worksheet"""
    #     # Start after title
    #     gantt_title_row = 3
    #     gantt_header_row = 4
    #     
    #     # Add Gantt section title
    #     gantt_title_cell = self.chart_worksheet.cell(row=gantt_title_row, column=1, value="Gantt Chart Data")
    #     gantt_title_cell.font = Font(size=12, bold=True)
    #     
    #     headers = ["Activity", "Start Time", "Duration", "Section", "Critical Path"]
    #     
    #     # Add headers
    #     for col, header in enumerate(headers, 1):
    #         cell = self.chart_worksheet.cell(row=gantt_header_row, column=col, value=header)
    #         cell.font = Font(bold=True)
    #         cell.border = self.border
    #         if col == 5:  # Critical Path column
    #             cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    #     
    #     # Add all activity data
    #     for row, activity in enumerate(chart_data['activities'], gantt_header_row + 1):
    #         # Activity name
    #         name_cell = self.chart_worksheet.cell(row=row, column=1, value=activity['name'])
    #         name_cell.border = self.border
    #         
    #         # Start time
    #         start_cell = self.chart_worksheet.cell(row=row, column=2, value=activity['start_time'])
    #         start_cell.border = self.border
    #         
    #         # Duration
    #         duration_cell = self.chart_worksheet.cell(row=row, column=3, value=activity['duration'])
    #         duration_cell.border = self.border
    #         
    #         # Section
    #         section_cell = self.chart_worksheet.cell(row=row, column=4, value=activity['section'])
    #         section_cell.border = self.border
    #         
    #         # Critical path indicator
    #         critical_cell = self.chart_worksheet.cell(row=row, column=5, value="YES" if activity['is_critical'] else "NO")
    #         critical_cell.border = self.border
    #         if activity['is_critical']:
    #             critical_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    #             critical_cell.font = Font(color="FFFFFF", bold=True)
    #             # Also highlight the activity name in red
    #             name_cell.font = Font(color="FF0000", bold=True)

    # def _create_s_curve(self, chart_data, start_row: int) -> int:
    #     """Create an S-curve showing ideal project progress over time"""
    #     try:
    #         from openpyxl.chart import LineChart, Reference
    #         
    #         # Add title on main sheet
    #         title_cell = self.worksheet.cell(row=start_row, column=1, value="Project S-Curve - Ideal Progress Over Time")
    #         title_cell.font = Font(size=14, bold=True)
    #         
    #         # Create data table on chart worksheet
    #         self._create_s_curve_data_table(chart_data)
    #         
    #         # Create chart on main sheet but reference data from chart worksheet
    #         chart = LineChart()
    #         chart.title = "Project S-Curve - Ideal Progress Timeline"
    #         chart.style = 13
    #         chart.y_axis.title = "Progress Percentage (%)"
    #         chart.x_axis.title = "Project Timeline (Days)"
    #         
    #         # Find the start row for S-curve data on chart worksheet (after Gantt data)
    #         s_curve_start_row = 4 + len(chart_data['activities']) + 3  # Gap after Gantt data
    #         
    #         # Reference data from chart worksheet
    #         activity_data = Reference(self.chart_worksheet, min_col=2, min_row=s_curve_start_row,
    #                                 max_col=2, max_row=s_curve_start_row + len(chart_data['progress_timeline']))
    #         duration_data = Reference(self.chart_worksheet, min_col=3, min_row=s_curve_start_row,
    #                                 max_col=3, max_row=s_curve_start_row + len(chart_data['progress_timeline']))
    #         timeline = Reference(self.chart_worksheet, min_col=1, min_row=s_curve_start_row + 1,
    #                            max_col=1, max_row=s_curve_start_row + len(chart_data['progress_timeline']))
    #         
    #         # Add both progress curves
    #         chart.add_data(activity_data, titles_from_data=True)
    #         chart.add_data(duration_data, titles_from_data=True)
    #         chart.set_categories(timeline)
    #         
    #         # Style the lines
    #         if len(chart.series) >= 1:
    #             s1 = chart.series[0]  # Activity completion progress
    #             s1.marker.symbol = "circle"
    #             s1.marker.size = 4
    #             s1.smooth = True
    #             s1.graphicalProperties.line.solidFill = "0066CC"  # Blue
    #         
    #         if len(chart.series) >= 2:
    #             s2 = chart.series[1]  # Duration progress
    #             s2.marker.symbol = "triangle"
    #             s2.marker.size = 4
    #             s2.smooth = True
    #             s2.graphicalProperties.line.solidFill = "FF6600"  # Orange
    #         
    #         # Position the chart
    #         chart.anchor = f"A{start_row + 2}"
    #         chart.width = 25
    #         chart.height = 12
    #         
    #         self.worksheet.add_chart(chart)
    #         
    #         return start_row + 15
    #         
    #     except Exception as e:
    #         print(f"Warning: Could not create S-curve: {e}")
    #         return start_row + 10

    # def _create_s_curve_data_table(self, chart_data):
    #     """Create S-curve data table on the chart worksheet"""
    #     # Start after Gantt data
    #     s_curve_title_row = 4 + len(chart_data['activities']) + 2
    #     start_row = 4 + len(chart_data['activities']) + 3
    #     
    #     # Add S-curve section title
    #     s_curve_title_cell = self.chart_worksheet.cell(row=s_curve_title_row, column=1, value="S-Curve Data")
    #     s_curve_title_cell.font = Font(size=12, bold=True)
    #     
    #     headers = ["Timeline (Days)", "Activity Progress (%)", "Duration Progress (%)", "Cumulative Duration"]
    #     
    #     # Add headers
    #     for col, header in enumerate(headers, 1):
    #         cell = self.chart_worksheet.cell(row=start_row, column=col, value=header)
    #         cell.font = Font(bold=True)
    #         cell.border = self.border
    #         if col == 2:  # Activity Progress column
    #             cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    #         elif col == 3:  # Duration Progress column
    #             cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    #     
    #     # Add progress timeline data
    #     for row, item in enumerate(chart_data['progress_timeline'], start_row + 1):
    #         # Timeline (in days)
    #         time_cell = self.chart_worksheet.cell(row=row, column=1, value=item['time'])
    #         time_cell.border = self.border
    #         
    #         # Activity progress percentage
    #         activity_cell = self.chart_worksheet.cell(row=row, column=2, value=round(item['activity_progress'], 1))
    #         activity_cell.border = self.border
    #         activity_cell.number_format = '0.0'
    #         
    #         # Duration progress percentage
    #         duration_cell = self.chart_worksheet.cell(row=row, column=3, value=round(item['duration_progress'], 1))
    #         duration_cell.border = self.border
    #         duration_cell.number_format = '0.0'
    #         
    #         # Cumulative duration
    #         cumulative_cell = self.chart_worksheet.cell(row=row, column=4, value=item['cumulative_duration'])
         #         cumulative_cell.border = self.border

    # END OF CHART GENERATION METHODS - ALL DISABLED


class GanttChartGenerator:
    """Generates Gantt chart worksheet based on Agile Gantt chart template"""
    
    def __init__(self, project: Project, workbook: openpyxl.Workbook):
        self.project = project
        self.workbook = workbook
        
        # Create Gantt chart worksheet
        self.gantt_worksheet = self.workbook.create_sheet("Gantt Chart")
        
        # Styling constants matching Agile Gantt chart
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Task type colors (matching project scheduler logic)
        self.task_type_map = {
            ActivitySection.PRE_KICKOFF: "Milestone",  # Pre-kickoff as milestones
            ActivitySection.POST_KICKOFF: "Goal"       # Post-kickoff as goals
        }
    
    def generate_gantt_chart(self) -> None:
        """Generate the complete Gantt chart worksheet"""
        # Calculate project timeline and dates
        # Use project's start date if available, otherwise default to today
        project_start_date = self.project.start_date if self.project.start_date else datetime.date.today()
        timeline_data = self._calculate_timeline_data(project_start_date)
        
        # Set up worksheet structure
        self._setup_worksheet_structure(timeline_data)
        
        # Add headers and timeline
        self._add_gantt_headers(timeline_data)
        
        # Add task data and visualization
        self._add_task_data(timeline_data)
        
        # Apply Gantt visualization formulas
        self._apply_gantt_formulas(timeline_data)
        
        # Apply direct cell styling instead of conditional formatting to prevent corruption
        self._apply_direct_gantt_styling(timeline_data)
        
        # Apply final formatting
        self._apply_final_formatting(timeline_data)
    
    def _calculate_timeline_data(self, start_date: datetime.date) -> Dict:
        """Calculate timeline dates and task schedules"""
        schedules = ScheduleCalculator.calculate_schedules(self.project)
        
        # Get critical path activities
        max_duration_activities = ScheduleCalculator.get_max_duration_activities(self.project)
        critical_activity_ids = {id(activity) for activity, _ in max_duration_activities}
        
        # Pre-calculate start working days for post-kickoff sequences
        post_sequences = sorted(self.project.get_sequences_by_section(ActivitySection.POST_KICKOFF))
        seq_start_working_days = {}
        cumulative = 0
        for seq in post_sequences:
            seq_start_working_days[seq] = cumulative
            seq_activities = [a for a in self.project.activities if a.section == ActivitySection.POST_KICKOFF and a.sequence == seq]
            max_duration = max(a.duration for a in seq_activities) if seq_activities else 0
            cumulative += max_duration

        # Convert schedules to actual dates
        task_timeline = []
        project_start_day = 0
        project_end_day = 0
        
        for activity in self.project.activities:
            if activity.section == ActivitySection.PRE_KICKOFF:
                # Pre-kickoff activities end at 0, and start backwards from 0
                task_start_working_day = -activity.duration
                task_start_day = ScheduleCalculator.apply_calendar_format(task_start_working_day, self.project.calendar_format)
                task_end_day = 0
                project_start_day = min(project_start_day, task_start_day)
            else:
                # Post-kickoff activities are left-aligned to the start of their sequence
                start_working_day = seq_start_working_days.get(activity.sequence, 0)
                end_working_day = start_working_day + activity.duration
                
                task_start_day = ScheduleCalculator.apply_calendar_format(start_working_day, self.project.calendar_format)
                task_end_day = ScheduleCalculator.apply_calendar_format(end_working_day, self.project.calendar_format)
            
            task_start_date = start_date + datetime.timedelta(days=task_start_day)
            task_end_date = start_date + datetime.timedelta(days=task_end_day)
            
            # Determine if this activity is on the critical path
            is_critical = id(activity) in critical_activity_ids
            
            task_timeline.append({
                'activity': activity,
                'start_date': task_start_date,
                'end_date': task_end_date,
                'start_day': task_start_day,
                'end_day': task_end_day,
                'duration': activity.duration,
                'task_type': self.task_type_map[activity.section],
                'is_critical': is_critical
            })
            
            project_end_day = max(project_end_day, task_end_day)
        
        # Calculate dynamic timeline to ensure ALL activities are captured
        timeline_buffer = 15  # 15 days buffer on each side
        timeline_start_day = project_start_day - timeline_buffer
        timeline_end_day = project_end_day + timeline_buffer
        timeline_days = timeline_end_day - timeline_start_day + 1
        
        # Ensure minimum 80 days for readability and prevent invalid ranges
        timeline_days = max(timeline_days, 80)
        
        # Additional safety check to prevent corruption
        if timeline_days <= 0:
            print("Warning: Invalid timeline calculation, using default 80 days")
            timeline_days = 80
        
        timeline_start = start_date + datetime.timedelta(days=timeline_start_day)
        date_timeline = []
        
        for i in range(timeline_days):
            date_timeline.append(timeline_start + datetime.timedelta(days=i))
        
        return {
            'project_start_date': start_date,
            'date_timeline': date_timeline,
            'task_timeline': task_timeline,
            'timeline_start': timeline_start,
            'timeline_days': timeline_days,
            'timeline_start_day': timeline_start_day,
            'critical_activities': critical_activity_ids
        }
    
    def _setup_worksheet_structure(self, timeline_data: Dict) -> None:
        """Set up basic worksheet structure and column widths"""
        # Set column widths (matching Agile Gantt chart)
        self.gantt_worksheet.column_dimensions['A'].width = 4   # Row numbers
        self.gantt_worksheet.column_dimensions['B'].width = 25  # Task names
        self.gantt_worksheet.column_dimensions['C'].width = 12  # Task types
        self.gantt_worksheet.column_dimensions['D'].width = 8   # Progress
        self.gantt_worksheet.column_dimensions['E'].width = 10  # Start dates
        self.gantt_worksheet.column_dimensions['F'].width = 12  # Start (calculated)
        self.gantt_worksheet.column_dimensions['G'].width = 8   # Duration
        self.gantt_worksheet.column_dimensions['H'].width = 5   # Empty
        
        # Timeline columns (I onwards) - dynamic based on project length
        timeline_columns = timeline_data['timeline_days']
        max_col = 9 + timeline_columns
        
        for col in range(9, max_col):  # Dynamic range based on timeline
            col_letter = get_column_letter(col)
            self.gantt_worksheet.column_dimensions[col_letter].width = 2.5
    
    def _add_gantt_headers(self, timeline_data: Dict) -> None:
        """Add headers matching Agile Gantt chart structure"""
        ws = self.gantt_worksheet
        
        # Row 1: Project title header
        ws.merge_cells('A1:H1')
        title_cell = ws.cell(row=1, column=1, value=f"Project Gantt Chart - {self.project.title}")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Row 2: Empty spacer
        
        # Row 3: Legend
        ws.cell(row=3, column=7, value="Legend:")
        
        # Row 4: Legend items (merge 2 cells each for better visibility)
        ws.merge_cells('L4:M4')
        goal_cell = ws.cell(row=4, column=12, value="G")
        goal_cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        goal_cell.font = Font(color="FFFFFF", bold=True)
        goal_cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('N4:O4')
        milestone_cell = ws.cell(row=4, column=14, value="M")
        milestone_cell.fill = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid")
        milestone_cell.font = Font(color="FFFFFF", bold=True)
        milestone_cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('P4:Q4')
        critical_cell = ws.cell(row=4, column=16, value="CP")
        critical_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
        critical_cell.font = Font(color="FFFFFF", bold=True)
        critical_cell.alignment = Alignment(horizontal='center')
        
        # Row 5: Project start date
        ws.cell(row=5, column=2, value="Project start date:")
        start_date_cell = ws.cell(row=5, column=6, value=timeline_data['project_start_date'])
        start_date_cell.number_format = 'dd-mmm-yyyy'
        
        # Row 6: Month headers
        self._add_month_headers(timeline_data)
        
        # Row 7: Date headers
        self._add_date_headers(timeline_data)
        
        # Row 8: Empty spacer
        
        # Row 9: Column headers
        headers = ["#", "Milestone description", "Type", "%", "Due date", "Start", "Days", "", "Timeline →"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col <= 7:  # Main headers
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = self.border
    
    def _add_month_headers(self, timeline_data: Dict) -> None:
        """Add month headers (row 6) like Agile Gantt chart"""
        ws = self.gantt_worksheet
        
        # Add month headers starting from column I (9)
        timeline_start_col = 9
        current_month = None
        month_start_col = None
        max_timeline_col = timeline_start_col + timeline_data['timeline_days'] - 1
        
        for i, date in enumerate(timeline_data['date_timeline']):
            col = timeline_start_col + i
            if col > max_timeline_col:  # Dynamic limit based on timeline
                break
                
            month_name = date.strftime("%B")
            
            if month_name != current_month:
                # End previous month merge if exists
                if current_month and month_start_col:
                    if col - 1 > month_start_col:
                        ws.merge_cells(f"{get_column_letter(month_start_col)}6:{get_column_letter(col-1)}6")
                    month_cell = ws.cell(row=6, column=month_start_col, value=current_month)
                    month_cell.font = Font(bold=True, size=10)
                    month_cell.alignment = Alignment(horizontal='center')
                    month_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Start new month
                current_month = month_name
                month_start_col = col
        
        # Handle last month
        if current_month and month_start_col:
            last_col = min(max_timeline_col, timeline_start_col + len(timeline_data['date_timeline']) - 1)
            if last_col > month_start_col:
                ws.merge_cells(f"{get_column_letter(month_start_col)}6:{get_column_letter(last_col)}6")
            month_cell = ws.cell(row=6, column=month_start_col, value=current_month)
            month_cell.font = Font(bold=True, size=10)
            month_cell.alignment = Alignment(horizontal='center')
            month_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    def _add_date_headers(self, timeline_data: Dict) -> None:
        """Add date headers (row 7) for timeline"""
        ws = self.gantt_worksheet
        
        timeline_start_col = 9
        max_timeline_col = timeline_start_col + timeline_data['timeline_days'] - 1
        
        for i, date in enumerate(timeline_data['date_timeline']):
            col = timeline_start_col + i
            if col > max_timeline_col:  # Dynamic limit based on timeline
                break
                
            date_cell = ws.cell(row=7, column=col, value=date)
            date_cell.number_format = 'd'  # Day number only
            date_cell.font = Font(size=8)
            date_cell.alignment = Alignment(horizontal='center')
    
    def _add_task_data(self, timeline_data: Dict) -> None:
        """Add task data rows"""
        ws = self.gantt_worksheet
        
        start_row = 10  # Start after headers
        for i, task_data in enumerate(timeline_data['task_timeline']):
            row = start_row + i
            activity = task_data['activity']
            
            # Column A: Row number
            ws.cell(row=row, column=1, value=i + 1)
            
            # Column B: Task name
            task_cell = ws.cell(row=row, column=2, value=activity.task)
            task_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Column C: Task type
            type_cell = ws.cell(row=row, column=3, value=task_data['task_type'])
            
            # Column D: Progress (empty for now)
            ws.cell(row=row, column=4, value="")
            
            # Column E: Due date (end date)
            due_cell = ws.cell(row=row, column=5, value=task_data['end_date'])
            due_cell.number_format = 'dd-mmm'
            
            # Column F: Start date
            start_cell = ws.cell(row=row, column=6, value=task_data['start_date'])
            start_cell.number_format = 'dd-mmm'
            
            # Column G: Duration
            ws.cell(row=row, column=7, value=activity.duration)
            
            # Apply basic formatting
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col == 1:  # Row number
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=9)
    
    def _apply_gantt_formulas(self, timeline_data: Dict) -> None:
        """Apply Gantt visualization formulas like Agile Gantt chart"""
        ws = self.gantt_worksheet
        
        start_row = 10
        timeline_start_col = 9
        max_timeline_col = timeline_start_col + timeline_data['timeline_days'] - 1
        
        for i, task_data in enumerate(timeline_data['task_timeline']):
            row = start_row + i
            
            # Add formulas for each timeline column
            for j, timeline_date in enumerate(timeline_data['date_timeline']):
                col = timeline_start_col + j
                if col > max_timeline_col:  # Dynamic limit based on timeline
                    break
                
                col_letter = get_column_letter(col)
                
                # Create formula that returns numeric values for conditional formatting
                # Use numbers instead of text for completely clean bars (no visible text)
                # 4 = Critical path goals (red), 3 = Critical milestones (dark red)
                # 2 = Regular goals (blue), 1 = Regular milestones (solid yellow)
                
                if task_data['is_critical']:
                    # Critical path activities (red bars) - return 4 or 3
                    if task_data['task_type'] == "Goal":
                        formula = (f'=IF(AND($C{row}="Goal",{col_letter}$7>=$F{row},{col_letter}$7<=$F{row}+$G{row}-1),4,"")') 
                    else:
                        formula = (f'=IF(AND($C{row}="Milestone",{col_letter}$7>=$F{row},{col_letter}$7<=$F{row}+$G{row}-1),3,"")') 
                else:
                    # Regular activities - return 2 or 1
                    formula = (f'=IF(AND($C{row}="Goal",{col_letter}$7>=$F{row},{col_letter}$7<=$F{row}+$G{row}-1),2,'
                              f'IF(AND($C{row}="Milestone",{col_letter}$7>=$F{row},{col_letter}$7<=$F{row}+$G{row}-1),1,""))')
                
                cell = ws.cell(row=row, column=col, value=formula)
                # Hide the numbers by setting custom number format that shows nothing
                cell.number_format = ';;;"";'  # This format hides all values (positive, negative, zero, text)
    
    def _apply_direct_gantt_styling(self, timeline_data: Dict) -> None:
        """Apply direct cell styling for Gantt visualization (avoiding conditional formatting corruption)"""
        ws = self.gantt_worksheet
        
        # Validate timeline data
        if timeline_data['timeline_days'] <= 0 or len(self.project.activities) == 0:
            print("Warning: Invalid timeline data, skipping Gantt styling")
            return
        
        timeline_start_col_num = 9
        timeline_end_col_num = timeline_start_col_num + timeline_data['timeline_days'] - 1
        timeline_end_row = 10 + len(self.project.activities) - 1
        
        # Safety checks
        if timeline_end_col_num > 16384:
            timeline_end_col_num = 16384
        if timeline_end_row < 10:
            return
        
        # Define color fills
        color_map = {
            4: PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid"),  # Critical goals - red
            3: PatternFill(start_color="B22222", end_color="B22222", fill_type="solid"),  # Critical milestones - dark red  
            2: PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),  # Regular goals - blue
            1: PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid"),  # Regular milestones - solid yellow
        }
        
        print(f"Applying direct styling to range: I10 to {get_column_letter(timeline_end_col_num)}{timeline_end_row}")
        
        try:
            # Apply styling directly to cells based on their calculated values
            for row in range(10, timeline_end_row + 1):
                for col in range(timeline_start_col_num, timeline_end_col_num + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Evaluate the formula to get the expected result
                        try:
                            # Get task data for this row
                            task_index = row - 10
                            if task_index < len(timeline_data['task_timeline']):
                                task_data = timeline_data['task_timeline'][task_index]
                                
                                # Get the date for this column
                                date_index = col - timeline_start_col_num
                                if date_index < len(timeline_data['date_timeline']):
                                    current_date = timeline_data['date_timeline'][date_index]
                                    task_start_date = task_data['start_date']
                                    task_end_date = task_start_date + datetime.timedelta(days=task_data['duration'] - 1)
                                    
                                    # Check if this date falls within the task duration
                                    if task_start_date <= current_date <= task_end_date:
                                        # Determine the color based on task type and criticality
                                        if task_data['is_critical']:
                                            if task_data['task_type'] == "Goal":
                                                color_value = 4  # Critical goal - red
                                            else:
                                                color_value = 3  # Critical milestone - dark red
                                        else:
                                            if task_data['task_type'] == "Goal":
                                                color_value = 2  # Regular goal - blue
                                            else:
                                                color_value = 1  # Regular milestone - green
                                        
                                        # Apply the color
                                        if color_value in color_map:
                                            cell.fill = color_map[color_value]
                                            # Make font invisible by setting it to white
                                            cell.font = Font(color="FFFFFF", size=1)
                                            # Keep the number format to hide values
                                            cell.number_format = ';;;"";'
                        except Exception as cell_error:
                            # Skip problematic cells
                            pass
            
            print("Direct Gantt styling applied successfully")
            
        except Exception as e:
            print(f"Warning: Could not apply direct Gantt styling: {e}")
            # Continue without styling rather than corrupting the file
    
    def _apply_final_formatting(self, timeline_data: Dict) -> None:
        """Apply final formatting touches"""
        ws = self.gantt_worksheet
        
        # Set row heights
        for row in range(1, 10 + len(self.project.activities)):
            if row == 1:  # Title row
                ws.row_dimensions[row].height = 25
            elif row in [6, 7, 9]:  # Header rows
                ws.row_dimensions[row].height = 20
            elif row >= 10:  # Activity data rows - increased height for better visibility
                ws.row_dimensions[row].height = 50
            else:  # Other rows (spacers, etc.)
                ws.row_dimensions[row].height = 18
        
        # Freeze panes to keep headers visible
        ws.freeze_panes = 'I10'
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False


class ExcelLoader:
    """Loads existing Excel files generated by the project scheduler"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def load_project(self) -> Optional[Project]:
        """Load project data from Excel file"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            # Extract project data
            project_title = self._extract_project_title()
            calendar_format = self._extract_calendar_format()
            activities = self._extract_activities()
            
            if not project_title:
                raise ValueError("Could not find project title in the Excel file")
            
            return Project(
                title=project_title,
                calendar_format=calendar_format,
                activities=activities
            )
            
        except Exception as e:
            raise Exception(f"Error loading Excel file: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _extract_project_title(self) -> Optional[str]:
        """Extract project title from the Excel file (supports both old and new formats)"""
        # First check for old format: "Project Title: ..." in column 1
        for row in range(1, 10):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str) and "Project Title:" in cell.value:
                return cell.value.replace("Project Title:", "").strip()
        
        # Then check for new format: title is stored in column 3 (C) in the generated Excel files
        for row in range(1, 10):
            title_cell = self.worksheet.cell(row=row, column=3)
            if title_cell.value and isinstance(title_cell.value, str) and title_cell.value.strip():
                # Skip if it's clearly a header row
                if "Activities/Tasks" in title_cell.value or "S/No" in title_cell.value:
                    continue
                return title_cell.value.strip()
        
        return None
    
    def _extract_calendar_format(self) -> CalendarFormat:
        """Extract calendar format from schedule calculations or default to 6-day"""
        # For now, default to 6-day format
        # Could be enhanced to detect from formulas in comments
        return CalendarFormat.SIX_DAY
    
    def _extract_activities(self) -> List[Activity]:
        """Extract activities from the Excel sheet"""
        activities = []
        current_section = None
        header_row = None
        
        # Find the header row - support both new format (S/No) and old format (Activities/Tasks)
        header_row = None
        old_format = False
        
        for row in range(1, 20):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value == "S/No":
                header_row = row
                old_format = False
                break
            elif cell.value == "Activities/Tasks":
                header_row = row
                old_format = True
                break
        
        if not header_row:
            raise ValueError("Could not find header row with S/No or Activities/Tasks column")
        
        # Process rows after header
        for row in range(header_row + 1, self.worksheet.max_row + 1):
            row_data = self._get_row_data(row)
            
            if not any(row_data):  # Skip empty rows
                continue
            
            # Check if this is a section header
            if self._is_section_row(row):
                section_name = self.worksheet.cell(row=row, column=1).value
                if "Pre-Kickoff" in section_name or "Pre Kickoff" in section_name:
                    current_section = ActivitySection.PRE_KICKOFF
                elif "Post Kick-off" in section_name or "Post Kickoff" in section_name:
                    current_section = ActivitySection.POST_KICKOFF
                continue
            
            # Check if this is the total row
            if self._is_total_row(row):
                break
            
            # Extract activity data
            try:
                if old_format:
                    # Old format: Activities/Tasks | Action Needed | Duration | Precursor | Sequence | Schedule | Resources | Budget
                    task = row_data[0] if row_data[0] else ""
                    action_needed = row_data[1] if row_data[1] else ""
                    duration = int(row_data[2]) if row_data[2] and str(row_data[2]).replace('.', '').isdigit() else 0
                    precursor = row_data[3] if row_data[3] else ""
                    sequence = int(row_data[4]) if row_data[4] and str(row_data[4]).replace('.', '').isdigit() else 0
                    resources = row_data[6] if len(row_data) > 6 and row_data[6] else ""
                    
                    # Convert budget from base currency to millions (old format stores in base currency)
                    try:
                        budget_raw = float(row_data[7]) if len(row_data) > 7 and row_data[7] is not None else 0.0
                        budget = budget_raw  # Already in base currency
                    except (ValueError, TypeError):
                        budget = 0.0
                else:
                    # New format: S/No | Activities/Tasks | Action Needed | Duration | Precursor | Sequence | Schedule | Resources | Budget (MILLION)
                    s_n = row_data[0] if row_data[0] else ""
                    task = row_data[1] if row_data[1] else ""
                    action_needed = row_data[2] if row_data[2] else ""
                    duration = int(row_data[3]) if row_data[3] and str(row_data[3]).replace('.', '').isdigit() else 0
                    precursor = row_data[4] if row_data[4] else ""
                    sequence = int(row_data[5]) if row_data[5] and str(row_data[5]).replace('.', '').isdigit() else 0
                    resources = row_data[7] if len(row_data) > 7 and row_data[7] else ""
                    
                    # Convert budget from millions back to base currency
                    try:
                        budget_millions = float(row_data[8]) if len(row_data) > 8 and row_data[8] is not None else 0.0
                        budget = budget_millions * 1000000
                    except (ValueError, TypeError):
                        budget = 0.0
                
                if task and current_section:  # Only add if we have a task and section
                    activity = Activity(
                        task=task,
                        action_needed=action_needed,
                        duration=duration,
                        precursor=precursor,
                        sequence=sequence,
                        resources=resources,
                        budget=budget,
                        section=current_section
                    )
                    activities.append(activity)
                    
            except (ValueError, TypeError) as e:
                # Skip rows with invalid data
                continue
        
        return activities
    
    def _get_row_data(self, row: int) -> List:
        """Get all cell values from a row"""
        return [self.worksheet.cell(row=row, column=col).value for col in range(1, 10)]
    
    def _is_section_row(self, row: int) -> bool:
        """Check if row is a section header"""
        cell = self.worksheet.cell(row=row, column=1)
        if not cell.value:
            return False
        
        value = str(cell.value).lower()
        return ("kickoff" in value or "kick-off" in value) and self._is_merged_across_columns(row, 1)
    
    def _is_total_row(self, row: int) -> bool:
        """Check if row is the total budget row"""
        # Check if "Total:" appears in Resources column (column 8)
        total_cell = self.worksheet.cell(row=row, column=8)
        return total_cell.value and str(total_cell.value).strip().lower() == "total:"
    
    def _is_merged_across_columns(self, row: int, col: int) -> bool:
        """Check if cell is merged across multiple columns"""
        cell = self.worksheet.cell(row=row, column=col)
        for merged_range in self.worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False


